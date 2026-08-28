"""Excel row probing for materialization (10,000 hard limit + 10,001 detection)."""

from __future__ import annotations

from typing import Any
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook

from apps.operation_analysis.services.datasource_preview.base import ConnectorError
from apps.operation_analysis.services.datasource_preview.excel import (
    MAX_EXCEL_BYTES,
    _normalize_dataframe_rows,
)

MAX_MATERIALIZE_ROWS = 10_000
MAX_XLSX_ARCHIVE_ENTRIES = 1_000
MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024


def validate_xlsx_archive_limits(file_obj) -> None:
    """Reject oversized decoded XLSX content before pandas/openpyxl allocation."""
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        with ZipFile(file_obj) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise ConnectorError("Excel 压缩包条目过多", code="excel_archive_too_large", status_code=400)
            decoded_size = sum(max(0, info.file_size) for info in infos)
            if decoded_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ConnectorError("Excel 解压后体积过大", code="excel_archive_too_large", status_code=400)
    except ConnectorError:
        raise
    except (BadZipFile, OSError, ValueError) as exc:
        raise ConnectorError("Excel 文件格式无效", code="excel_file_type_invalid", status_code=400) from exc
    finally:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)


def validate_xlsx_data_row_limit(file_obj, sheet_name: str | None = None) -> None:
    """Count non-empty data rows so blank physical rows cannot hide row 10,001."""
    workbook = None
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)  # Header row follows pandas.read_excel semantics.
        data_rows = 0
        for row in rows:
            if any(value is not None for value in row):
                data_rows += 1
                if data_rows > MAX_MATERIALIZE_ROWS:
                    raise ConnectorError(
                        f"Excel 行数超过 {MAX_MATERIALIZE_ROWS}，无法处理",
                        code="excel_rows_too_many",
                        status_code=400,
                    )
    except ConnectorError:
        raise
    except Exception as exc:
        raise ConnectorError(f"Excel 解析失败: {exc}", code="excel_parse_failed", status_code=400) from exc
    finally:
        if workbook is not None:
            workbook.close()
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)


def read_excel_rows_for_materialize(file_obj, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Parse .xlsx and fail closed if more than MAX_MATERIALIZE_ROWS data rows exist.

    Streams the sheet first so the 10,001st non-empty data row is detectable.
    """
    if not file_obj:
        raise ConnectorError("请上传 Excel 文件", code="excel_file_required", status_code=400)

    file_name = getattr(file_obj, "name", "") or ""
    if file_name and not file_name.lower().endswith(".xlsx"):
        raise ConnectorError("仅支持 Excel 文件（.xlsx）", code="excel_file_type_invalid", status_code=400)

    file_size = getattr(file_obj, "size", None)
    if file_size and file_size > MAX_EXCEL_BYTES:
        raise ConnectorError("Excel 文件不能超过 2MB", code="excel_file_too_large", status_code=400)

    validate_xlsx_archive_limits(file_obj)
    validate_xlsx_data_row_limit(file_obj, sheet_name=sheet_name)
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        dataframe = pd.read_excel(
            file_obj,
            sheet_name=sheet_name or 0,
        )
    except ConnectorError:
        raise
    except Exception as exc:
        raise ConnectorError(f"Excel 解析失败: {exc}", code="excel_parse_failed", status_code=400) from exc

    dataframe = dataframe.dropna(how="all")
    if dataframe.empty:
        raise ConnectorError("Excel 没有可处理的数据", code="excel_empty", status_code=400)

    if len(dataframe.index) > MAX_MATERIALIZE_ROWS:
        raise ConnectorError(
            f"Excel 行数超过 {MAX_MATERIALIZE_ROWS}，无法处理",
            code="excel_rows_too_many",
            status_code=400,
        )

    return _normalize_dataframe_rows(dataframe)
