"""Excel 双槽物化：行数探测、原子切换、失败保留、旧任务晚到。"""

from io import BytesIO
from itertools import chain
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest
from django.core.files.storage import InMemoryStorage
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.operation_analysis.models.datasource_models import DataSourceAPIModel
from apps.operation_analysis.models.excel_materialization_models import ExcelMaterializationSlot
from apps.operation_analysis.services.datasource_preview.base import ConnectorError
from apps.operation_analysis.services.excel_materialize import (
    ExcelMaterializer,
    MAX_MATERIALIZE_ROWS,
    load_excel_runtime,
    resolve_excel_runtime_status,
    submit_excel_candidate,
)
from apps.operation_analysis.services.excel_materialize import row_probe as row_probe_mod


@pytest.fixture(autouse=True)
def _local_excel_slot_storage():
    source_field = ExcelMaterializationSlot._meta.get_field("source_file")
    result_field = ExcelMaterializationSlot._meta.get_field("result_file")
    original_source = source_field.storage
    original_result = result_field.storage
    storage = InMemoryStorage(base_url="/test-media/")
    source_field.storage = storage
    result_field.storage = storage
    try:
        yield
    finally:
        source_field.storage = original_source
        result_field.storage = original_result


def _xlsx_bytes(rows: list[list], headers: list[str] | None = None) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers or ["name", "value"])
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _uploaded(rows: list[list], name: str = "demo.xlsx") -> SimpleUploadedFile:
    return SimpleUploadedFile(
        name=name,
        content=_xlsx_bytes(rows),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _excel_datasource(**kwargs) -> DataSourceAPIModel:
    defaults = {
        "name": kwargs.pop("name", f"excel-ds-{DataSourceAPIModel.objects.count() + 1}"),
        "rest_api": "",
        "source_type": DataSourceAPIModel.SOURCE_TYPE_EXCEL,
        "groups": [1],
        "query_config": {},
        "transform_config": {},
        "chart_type": ["table"],
    }
    defaults.update(kwargs)
    return DataSourceAPIModel.objects.create(**defaults)


def test_row_probe_rejects_10001st_data_row(monkeypatch):
    """中间空白物理行不能隐藏第 10001 条实际数据。"""
    worksheet = MagicMock()
    worksheet.iter_rows.return_value = chain(
        [("name",)],
        ((f"r{index}",) for index in range(5_000)),
        [(None,)],
        ((f"r{index}",) for index in range(5_000, MAX_MATERIALIZE_ROWS + 1)),
    )
    workbook = MagicMock()
    workbook.active = worksheet
    monkeypatch.setattr(row_probe_mod, "load_workbook", MagicMock(return_value=workbook))
    file_obj = MagicMock()
    file_obj.name = "big.xlsx"
    file_obj.size = 1024

    with pytest.raises(ConnectorError) as exc:
        row_probe_mod.validate_xlsx_data_row_limit(file_obj)

    assert exc.value.code == "excel_rows_too_many"


def test_row_probe_rejects_large_decoded_archive(monkeypatch):
    info = SimpleNamespace(file_size=row_probe_mod.MAX_XLSX_UNCOMPRESSED_BYTES + 1)
    archive = MagicMock()
    archive.__enter__.return_value.infolist.return_value = [info]
    monkeypatch.setattr(row_probe_mod, "ZipFile", MagicMock(return_value=archive))
    file_obj = MagicMock(name="large-archive.xlsx")

    with pytest.raises(ConnectorError) as exc:
        row_probe_mod.validate_xlsx_archive_limits(file_obj)

    assert exc.value.code == "excel_archive_too_large"


@pytest.mark.django_db
def test_materialize_without_script_promotes_success_and_clears_imported_items():
    ds = _excel_datasource(
        query_config={
            "imported_items": [{"name": "旧", "value": 1}],
            "imported_fields": [{"key": "name", "title": "name", "value_type": "string"}],
        }
    )
    slot = submit_excel_candidate(
        ds,
        uploaded_file=_uploaded([["官网", 120], ["广告", 96]]),
        schedule=False,
    )

    result = ExcelMaterializer().materialize_candidate(slot.id)
    assert result["ok"] is True

    ds.refresh_from_db()
    assert ds.excel_success_slot_id == slot.id
    assert ds.excel_candidate_slot_id is None
    assert "imported_items" not in (ds.query_config or {})

    success = ds.excel_success_slot
    assert success.status == ExcelMaterializationSlot.STATUS_SUCCEEDED
    assert success.role == ExcelMaterializationSlot.ROLE_SUCCESS
    assert success.row_count == 2
    assert resolve_excel_runtime_status(ds) == "ready"

    runtime = load_excel_runtime(ds, limit=10)
    assert runtime["items"] == [
        {"name": "官网", "value": 120},
        {"name": "广告", "value": 96},
    ]
    assert runtime["warnings"] == []


@pytest.mark.django_db
def test_materialize_failure_keeps_previous_success(monkeypatch):
    ds = _excel_datasource()
    first = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    assert ExcelMaterializer().materialize_candidate(first.id)["ok"] is True
    ds.refresh_from_db()
    success_id = ds.excel_success_slot_id

    second = submit_excel_candidate(ds, uploaded_file=_uploaded([["b", 2]]), schedule=False)

    def boom(*args, **kwargs):
        raise ConnectorError("Excel 行数超过限制", code="excel_rows_too_many", status_code=400)

    monkeypatch.setattr(
        "apps.operation_analysis.services.excel_materialize.materializer.read_excel_rows_for_materialize",
        boom,
    )
    result = ExcelMaterializer().materialize_candidate(second.id)
    assert result["ok"] is False
    assert result["code"] == "excel_rows_too_many"

    ds.refresh_from_db()
    assert ds.excel_success_slot_id == success_id
    assert ds.excel_candidate_slot_id == second.id
    second.refresh_from_db()
    assert second.status == ExcelMaterializationSlot.STATUS_FAILED
    assert resolve_excel_runtime_status(ds) == "update_failed_using_previous"

    runtime = load_excel_runtime(ds, limit=10)
    assert runtime["items"] == [{"name": "a", "value": 1}]
    assert runtime["warnings"]


@pytest.mark.django_db
def test_stale_candidate_does_not_overwrite_newer_success():
    ds = _excel_datasource()
    first = submit_excel_candidate(ds, uploaded_file=_uploaded([["old", 1]]), schedule=False)
    assert ExcelMaterializer().materialize_candidate(first.id)["ok"] is True

    stale = submit_excel_candidate(ds, uploaded_file=_uploaded([["stale", 9]]), schedule=False)
    newer = submit_excel_candidate(ds, uploaded_file=_uploaded([["new", 2]]), schedule=False)
    ds.refresh_from_db()
    assert ds.excel_candidate_slot_id == newer.id
    assert not ExcelMaterializationSlot.objects.filter(pk=stale.id).exists()

    orphan = ExcelMaterializationSlot.objects.create(
        datasource=ds,
        role=ExcelMaterializationSlot.ROLE_CANDIDATE,
        generation=1,
        status=ExcelMaterializationSlot.STATUS_PENDING,
        source_filename="orphan.xlsx",
    )
    orphan.source_file.save("orphan.xlsx", _uploaded([["orphan", 8]]), save=True)
    assert ExcelMaterializer().materialize_candidate(orphan.id)["code"] == "slot_stale"

    assert ExcelMaterializer().materialize_candidate(newer.id)["ok"] is True
    ds.refresh_from_db()
    assert ds.excel_success_slot_id == newer.id
    runtime = load_excel_runtime(ds, limit=10)
    assert runtime["items"] == [{"name": "new", "value": 2}]


@pytest.mark.django_db
def test_legacy_imported_items_still_run_without_slots():
    ds = _excel_datasource(
        query_config={
            "imported_items": [{"name": "官网", "value": 120}],
            "imported_fields": [{"key": "name", "title": "name", "value_type": "string"}],
        }
    )
    assert resolve_excel_runtime_status(ds) == "ready"
    runtime = load_excel_runtime(ds, limit=10)
    assert runtime["items"] == [{"name": "官网", "value": 120}]


@pytest.mark.django_db
def test_needs_upload_when_no_legacy_and_no_slots():
    ds = _excel_datasource()
    assert resolve_excel_runtime_status(ds) == "needs_upload"
    with pytest.raises(ConnectorError) as exc:
        load_excel_runtime(ds, limit=10)
    assert exc.value.code == "excel_needs_upload"


@pytest.mark.django_db
def test_excel_can_retry_only_with_saved_source_and_failed():
    from apps.operation_analysis.services.excel_materialize import (
        build_excel_materialization_payload,
        excel_can_retry,
        excel_has_saved_source,
    )

    empty = _excel_datasource()
    assert excel_has_saved_source(empty) is False
    assert excel_can_retry(empty) is False
    payload = build_excel_materialization_payload(empty)
    assert payload["status"] == "needs_upload"
    assert payload["can_retry"] is False
    assert payload["has_saved_source"] is False

    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    ExcelMaterializationSlot.objects.filter(pk=slot.id).update(
        status=ExcelMaterializationSlot.STATUS_FAILED,
        error_code="excel_materialize_internal_error",
        error_summary="boom",
    )
    ds.refresh_from_db()
    assert excel_has_saved_source(ds) is True
    assert resolve_excel_runtime_status(ds) == "failed"
    assert excel_can_retry(ds) is True
    assert build_excel_materialization_payload(ds)["can_retry"] is True


@pytest.mark.django_db
def test_discard_unready_excel_datasource_deletes_shell():
    from apps.operation_analysis.services.excel_materialize import discard_unready_excel_datasource

    ds = _excel_datasource(name="to-discard")
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    ExcelMaterializationSlot.objects.filter(pk=slot.id).update(status=ExcelMaterializationSlot.STATUS_FAILED)
    ds_id = ds.id
    assert discard_unready_excel_datasource(ds) is True
    assert not DataSourceAPIModel.objects.filter(pk=ds_id).exists()
    assert not ExcelMaterializationSlot.objects.filter(datasource_id=ds_id).exists()


@pytest.mark.django_db
def test_deleting_slot_deletes_files_after_commit(django_capture_on_commit_callbacks):
    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    slot.result_file.save("result.json.gz", SimpleUploadedFile("result.json.gz", b"result"), save=True)
    source_storage = slot.source_file.storage
    result_storage = slot.result_file.storage
    source_name = slot.source_file.name
    result_name = slot.result_file.name

    with django_capture_on_commit_callbacks(execute=True):
        slot.delete()

    assert source_storage.exists(source_name) is False
    assert result_storage.exists(result_name) is False


@pytest.mark.django_db
def test_submit_cleans_uploaded_file_when_database_transaction_rolls_back(monkeypatch):
    ds = _excel_datasource()
    storage = ExcelMaterializationSlot._meta.get_field("source_file").storage
    delete_file = MagicMock(wraps=storage.delete)
    monkeypatch.setattr(storage, "delete", delete_file)
    slot_count = ExcelMaterializationSlot.objects.count()

    def fail_datasource_save(*args, **kwargs):
        raise RuntimeError("database write failed")

    monkeypatch.setattr(DataSourceAPIModel, "save", fail_datasource_save)

    with pytest.raises(RuntimeError, match="database write failed"):
        submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)

    assert ExcelMaterializationSlot.objects.count() == slot_count
    delete_file.assert_called_once()


@pytest.mark.django_db
def test_discard_unready_refuses_when_success_exists():
    from apps.operation_analysis.services.excel_materialize import discard_unready_excel_datasource

    ds = _excel_datasource()
    first = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    assert ExcelMaterializer().materialize_candidate(first.id)["ok"] is True
    ds.refresh_from_db()
    assert discard_unready_excel_datasource(ds) is False
    assert DataSourceAPIModel.objects.filter(pk=ds.id).exists()


@pytest.mark.django_db
def test_inline_materialize_after_submit_reaches_ready():
    from apps.operation_analysis.services.excel_materialize import materialize_candidate_inline

    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["x", 9]]), schedule=False)
    result = materialize_candidate_inline(slot.id)
    assert result["ok"] is True
    ds.refresh_from_db()
    assert resolve_excel_runtime_status(ds) == "ready"


@pytest.mark.django_db
def test_materialize_candidate_claim_prevents_duplicate_worker(monkeypatch):
    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["x", 9]]), schedule=False)
    ExcelMaterializationSlot.objects.filter(pk=slot.id).update(
        status=ExcelMaterializationSlot.STATUS_PROCESSING
    )
    parser = MagicMock()
    monkeypatch.setattr(
        "apps.operation_analysis.services.excel_materialize.materializer.read_excel_rows_for_materialize",
        parser,
    )

    result = ExcelMaterializer().materialize_candidate(slot.id)

    assert result["ok"] is False
    assert result["code"] == "slot_in_progress"
    parser.assert_not_called()


@pytest.mark.django_db
def test_late_failure_does_not_overwrite_successful_slot():
    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["x", 9]]), schedule=False)
    assert ExcelMaterializer().materialize_candidate(slot.id)["ok"] is True

    ExcelMaterializer()._mark_failed(slot, code="late_failure", summary="late")

    slot.refresh_from_db()
    assert slot.status == ExcelMaterializationSlot.STATUS_SUCCEEDED


def test_safe_materialize_error_summary_is_actionable():
    from apps.operation_analysis.services.excel_materialize.materializer import (
        safe_materialize_error_summary,
    )

    assert "存储服务无权限" in safe_materialize_error_summary(Exception("S3Error AccessDenied"))
    assert "存储空间未就绪" in safe_materialize_error_summary(Exception("NoSuchBucket"))
    assert "后台处理服务" in safe_materialize_error_summary(Exception("kombu broker down"))
    generic = safe_materialize_error_summary(Exception("weird boom password=secret"))
    assert "重新选择文件" in generic
    assert "secret" not in generic


@pytest.mark.django_db
def test_schedule_resubmit_enqueues_celery_task(monkeypatch):
    from apps.operation_analysis.services.excel_materialize.submit import (
        schedule_resubmit_excel_from_saved_source,
    )

    captured = {}

    class FakeTask:
        @staticmethod
        def delay(datasource_id):
            captured["datasource_id"] = datasource_id

    monkeypatch.setattr(
        "apps.operation_analysis.tasks.tasks.resubmit_excel_from_saved_source_task",
        FakeTask,
    )

    monkeypatch.setattr(
        "apps.operation_analysis.services.excel_materialize.submit.transaction.on_commit",
        lambda callback: callback(),
    )

    schedule_resubmit_excel_from_saved_source(42)
    assert captured["datasource_id"] == 42


@pytest.mark.django_db
def test_resubmit_task_loads_script_from_database(monkeypatch):
    from apps.operation_analysis.tasks.tasks import resubmit_excel_from_saved_source_task

    ds = _excel_datasource(
        transform_config={"enabled": True, "language": "python", "script": "def transform(rows, params): return rows"}
    )
    captured = {}

    def fake_submit(datasource, *, transform_config, schedule):
        captured["datasource_id"] = datasource.id
        captured["transform_config"] = transform_config
        captured["schedule"] = schedule
        return SimpleNamespace(id=99, generation=2)

    monkeypatch.setattr(
        "apps.operation_analysis.services.excel_materialize.submit_excel_candidate_from_saved_source",
        fake_submit,
    )

    result = resubmit_excel_from_saved_source_task(ds.id)

    assert result == {"ok": True, "slot_id": 99, "generation": 2}
    assert captured["datasource_id"] == ds.id
    assert captured["transform_config"] == ds.transform_config
    assert captured["schedule"] is True


@pytest.mark.django_db
def test_abandon_excel_materialization_deletes_slots_and_files(django_capture_on_commit_callbacks):
    from apps.operation_analysis.services.excel_materialize import abandon_excel_materialization

    ds = _excel_datasource(query_config={"sheet_name": "Sheet1", "imported_items": [{"a": 1}]})
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    storage = slot.source_file.storage
    source_name = slot.source_file.name
    ds_id = ds.id

    with django_capture_on_commit_callbacks(execute=True):
        result = abandon_excel_materialization(ds)

    ds.refresh_from_db()
    assert result["deleted_slots"] >= 1
    assert ds.excel_success_slot_id is None
    assert ds.excel_candidate_slot_id is None
    assert ds.excel_materialization_generation == 0
    assert "imported_items" not in (ds.query_config or {})
    assert "sheet_name" not in (ds.query_config or {})
    assert not ExcelMaterializationSlot.objects.filter(datasource_id=ds_id).exists()
    assert storage.exists(source_name) is False
    assert abandon_excel_materialization(ds)["deleted_slots"] == 0


@pytest.mark.django_db
def test_deleting_datasource_deletes_slot_files(django_capture_on_commit_callbacks):
    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    storage = slot.source_file.storage
    source_name = slot.source_file.name

    with django_capture_on_commit_callbacks(execute=True):
        ds.delete()

    assert storage.exists(source_name) is False


@pytest.mark.django_db
def test_materialize_refuses_after_source_type_leaves_excel():
    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    DataSourceAPIModel.objects.filter(pk=ds.id).update(source_type=DataSourceAPIModel.SOURCE_TYPE_POSTGRESQL)

    result = ExcelMaterializer().materialize_candidate(slot.id)

    assert result["ok"] is False
    assert result["code"] == "not_excel"
    assert ExcelMaterializationSlot.objects.filter(pk=slot.id).exists()


@pytest.mark.django_db
def test_sweep_abandons_slots_left_on_non_excel_datasource(django_capture_on_commit_callbacks):
    from apps.operation_analysis.services.excel_materialize import sweep_abandoned_excel_materializations

    ds = _excel_datasource()
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    storage = slot.source_file.storage
    source_name = slot.source_file.name
    DataSourceAPIModel.objects.filter(pk=ds.id).update(source_type=DataSourceAPIModel.SOURCE_TYPE_POSTGRESQL)

    with django_capture_on_commit_callbacks(execute=True):
        stats = sweep_abandoned_excel_materializations()

    ds.refresh_from_db()
    assert stats["cleaned"] == 1
    assert ds.excel_candidate_slot_id is None
    assert not ExcelMaterializationSlot.objects.filter(datasource_id=ds.id).exists()
    assert storage.exists(source_name) is False


@pytest.mark.django_db
def test_serializer_switch_from_excel_to_postgresql_abandons_slots(authenticated_user, django_capture_on_commit_callbacks):
    from rest_framework.test import APIRequestFactory, force_authenticate

    from apps.operation_analysis.serializers.datasource_serializers import DataSourceAPIModelSerializer

    ds = _excel_datasource(name="excel-1")
    slot = submit_excel_candidate(ds, uploaded_file=_uploaded([["a", 1]]), schedule=False)
    storage = slot.source_file.storage
    source_name = slot.source_file.name

    request = APIRequestFactory().put("/operation_analysis/api/data_source/1/", data={}, format="json")
    request.COOKIES["current_team"] = "1"
    request.COOKIES["include_children"] = "0"
    request.user = authenticated_user
    force_authenticate(request, user=authenticated_user)

    serializer = DataSourceAPIModelSerializer(
        ds,
        context={"request": request},
        data={
            "name": "excel-1",
            "rest_api": "",
            "source_type": "postgresql",
            "connection": None,
            "connection_config": {
                "host": "127.0.0.1",
                "port": 5432,
                "database": "bklite",
                "username": "bklite",
                "password": "secret",
            },
            "query_config": {"sql": "SELECT 1", "table": ""},
            "params": [],
            "chart_type": ["table"],
            "field_schema": [],
            "groups": [1],
            "namespaces": [],
            "tag": [],
        },
    )
    assert serializer.is_valid(), serializer.errors
    with django_capture_on_commit_callbacks(execute=True):
        updated = serializer.save()

    assert updated.source_type == "postgresql"
    assert updated.excel_candidate_slot_id is None
    assert updated.excel_success_slot_id is None
    assert not ExcelMaterializationSlot.objects.filter(datasource_id=updated.id).exists()
    assert storage.exists(source_name) is False
