import os

import openpyxl
import pytest


WINSPHERE_MODULE = pytest.importorskip(
    "apps.cmdb_enterprise.collect.winsphere",
    reason="WinSphere 仅随企业采集扩展交付",
)


XLSX = os.path.join(
    os.path.dirname(__file__),
    "..",
    "support-files",
    "model_config.xlsx",
)

MODEL_IDS = (
    "winsphere",
    "winsphere_host_pool",
    "winsphere_cluster",
    "winsphere_host",
    "winsphere_vm",
    "winsphere_storage_pool",
    "winsphere_vswitch",
    "winsphere_port_group",
)

EXPECTED_ASSOCIATIONS = {
    "asso-winsphere_host_pool": {
        ("winsphere_host_pool", "winsphere", "group", "n:1"),
    },
    "asso-winsphere_cluster": {
        ("winsphere_cluster", "winsphere_host_pool", "group", "n:1"),
    },
    "asso-winsphere_host": {
        ("winsphere_host", "winsphere_cluster", "group", "n:1"),
        ("winsphere_host", "winsphere_host_pool", "group", "n:1"),
    },
    "asso-winsphere_vm": {
        ("winsphere_vm", "winsphere_host", "run", "n:1"),
    },
    "asso-winsphere_storage_pool": {
        ("winsphere_storage_pool", "winsphere_host", "connect", "n:n"),
    },
    "asso-winsphere_vswitch": {
        ("winsphere_vswitch", "winsphere_host", "connect", "n:n"),
    },
    "asso-winsphere_port_group": {
        ("winsphere_port_group", "winsphere_vswitch", "group", "n:1"),
    },
}


def _records(sheet):
    headers = [cell.value for cell in sheet[2]]
    return [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=3, values_only=True)
        if any(value is not None for value in values)
    ]


def test_winsphere_models_and_relations_match_collection_contract():
    workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    classifications = {
        row["classification_id"]: row["classification_name"]
        for row in _records(workbook["classifications"])
    }
    models = {
        row["model_id"]: row
        for row in _records(workbook["models"])
    }

    assert classifications["winsphere"] == "云宏 WinSphere"
    assert set(MODEL_IDS) <= set(models)
    assert all(models[model_id]["classification_id"] == "winsphere" for model_id in MODEL_IDS)

    WinsphereCollectionPlugin = WINSPHERE_MODULE.WinsphereCollectionPlugin

    for model_id in MODEL_IDS:
        sheet_name = f"attr-{model_id}"
        assert sheet_name in workbook.sheetnames
        attrs = {
            row["attr_id"]: row
            for row in _records(workbook[sheet_name])
        }
        assert attrs["inst_name"]["is_only"] is True
        if model_id == "winsphere":
            assert attrs["resource_id"]["is_only"] is True
        assert attrs["organization"]["is_required"] is True
        assert {"auto_collect", "collect_time", "collect_task"} <= set(attrs)
        mapping = WinsphereCollectionPlugin.field_mappings[model_id]
        assert set(mapping) - {"assos"} <= set(attrs)
        for attr_id, converter in mapping.items():
            if attr_id in {"assos", "inst_name"}:
                continue
            if isinstance(converter, tuple):
                expected_type = {
                    "to_int": "int",
                    "to_float": "float",
                    "to_bool": "bool",
                }[converter[0].__name__]
            else:
                expected_type = "str"
            assert attrs[attr_id]["attr_type"] == expected_type

    for sheet_name, expected in EXPECTED_ASSOCIATIONS.items():
        assert sheet_name in workbook.sheetnames
        actual = {
            (
                row["src_model_id"],
                row["dst_model_id"],
                row["asst_id"],
                row["mapping"],
            )
            for row in _records(workbook[sheet_name])
        }
        assert actual == expected
