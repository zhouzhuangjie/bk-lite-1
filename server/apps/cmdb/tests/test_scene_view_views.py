"""场景视图个人 CRUD 与执行接口（不落库，避开 sqlite 全量 migrate）。"""

import json
import logging
from types import SimpleNamespace
from unittest.mock import MagicMock

import pytest
from django.db.models import Q
from rest_framework import status
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.cmdb.models.scene_view import SceneView
from apps.cmdb.serializers.scene_view import SceneViewSerializer
from apps.cmdb.views.scene_view import SceneViewViewSet

pytestmark = pytest.mark.unit


def _user(*, username="alice", is_superuser=True, permission=None, domain="domain.com", roles=None):
    return SimpleNamespace(
        username=username,
        is_superuser=is_superuser,
        is_authenticated=True,
        is_active=True,
        domain=domain,
        locale="zh-Hans",
        permission=permission if permission is not None else {"cmdb": {"asset_info-View"}},
        group_list=[{"id": 1, "name": "Default Team"}],
        roles=["admin"] if roles is None else roles,
    )


def _req(method, user, data=None, cookies=None, query=""):
    factory = APIRequestFactory()
    fn = getattr(factory, method)
    path = f"/cmdb/api/scene_views/{query}"
    request = fn(path, data=data, format="json") if data is not None else fn(path)
    if cookies:
        request.COOKIES.update(cookies)
    force_authenticate(request, user=user)
    return request


def _body(response):
    if hasattr(response, "render"):
        response.render()
        return json.loads(response.rendered_content)
    return json.loads(response.content)


def _payload(**overrides):
    data = {
        "name": "测试环境 · 计算资产",
        "model_ids": ["host", "physcial_server"],
        "tags": ["env:test"],
        "tag_match": "and",
        "visibility": "personal",
    }
    data.update(overrides)
    return data


def _scene(**overrides):
    data = {
        "id": 7,
        "name": "测试环境 · 计算资产",
        "visibility": SceneView.Visibility.PERSONAL,
        "organization": None,
        "model_ids": ["host", "physcial_server"],
        "tags": ["env:test"],
        "tag_match": SceneView.TagMatch.AND,
        "created_by": "alice",
        "created_at": "2026-08-27T00:00:00Z",
        "updated_by": "alice",
        "updated_at": "2026-08-27T00:00:00Z",
        "domain": "domain.com",
    }
    data.update(overrides)
    return SimpleNamespace(**data)


def test_serializer_rejects_non_personal_and_empty_scope():
    org = SceneViewSerializer(data=_payload(visibility="organization"))
    assert org.is_valid() is False
    assert "visibility" in org.errors

    empty_models = SceneViewSerializer(data=_payload(model_ids=[]))
    assert empty_models.is_valid() is False
    assert "model_ids" in empty_models.errors

    empty_tags = SceneViewSerializer(data=_payload(tags=[]))
    assert empty_tags.is_valid() is False
    assert "tags" in empty_tags.errors


def test_serializer_accepts_personal_and_or():
    serializer = SceneViewSerializer(data=_payload(tag_match="or"))
    assert serializer.is_valid(), serializer.errors
    assert serializer.validated_data["tag_match"] == "or"
    assert serializer.validated_data["visibility"] == "personal"


def test_get_queryset_uses_visible_query(monkeypatch):
    captured = {}

    class _QS(list):
        def order_by(self, *_args):
            return self

    def fake_filter(*args, **kwargs):
        captured["args"] = args
        captured["kwargs"] = kwargs
        return _QS()

    monkeypatch.setattr(SceneView.objects, "filter", fake_filter)
    view = SceneViewViewSet()
    view.request = SimpleNamespace(user=_user())
    view.get_queryset()
    assert captured["kwargs"] == {}
    assert len(captured["args"]) == 1
    assert isinstance(captured["args"][0], Q)


def test_other_user_retrieve_is_404(monkeypatch):
    from django.http import Http404

    monkeypatch.setattr(
        SceneViewViewSet,
        "get_object",
        MagicMock(side_effect=Http404),
    )
    response = SceneViewViewSet.as_view({"get": "retrieve"})(_req("get", _user(username="bob")), pk=1)
    assert response.status_code == status.HTTP_404_NOT_FOUND


def test_execute_uses_opener_identity_and_strips_graph_ids(monkeypatch):
    scene = _scene()
    captured = {}
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: scene)

    def fake_execute(**kwargs):
        captured.update(kwargs)
        return {
            "total": 1,
            "models": [
                {
                    "model_id": "host",
                    "count": 1,
                    "insts": [{"_id": 9, "_labels": ["inst"], "inst_uuid": "h1", "inst_name": "host-1"}],
                }
            ],
        }

    monkeypatch.setattr("apps.cmdb.views.scene_view.execute_scene_query", fake_execute)
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view._model_table_columns",
        lambda model_id, creator: [{"attr_id": "inst_name", "attr_name": "名称", "attr_type": "str"}],
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.CmdbRulesFormatUtil.format_user_groups_permissions",
        lambda request, model_id: {f"perm-{model_id}": True},
    )
    response = SceneViewViewSet.as_view({"post": "execute"})(
        _req("post", _user(), {"page": 1, "page_size": 20}),
        pk=7,
    )
    assert response.status_code == status.HTTP_200_OK
    data = _body(response)["data"]
    assert data["total"] == 1
    assert data["models"][0]["insts"] == [{"inst_uuid": "h1", "inst_name": "host-1"}]
    assert data["models"][0]["columns"] == [{"attr_id": "inst_name", "attr_name": "名称", "attr_type": "str"}]
    assert captured["creator"] == "alice"
    assert captured["model_ids"] == ["host", "physcial_server"]
    assert captured["tags"] == ["env:test"]
    assert captured["tag_match"] == "and"
    assert captured["permission_map_loader"]("host") == {"perm-host": True}
    assert captured["pagination"] == {}


def test_execute_forwards_per_model_pagination(monkeypatch):
    captured = {}
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: _scene())
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.execute_scene_query",
        lambda **kwargs: captured.update(kwargs) or {"total": 0, "models": []},
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.CmdbRulesFormatUtil.format_user_groups_permissions",
        lambda request, model_id: {},
    )
    response = SceneViewViewSet.as_view({"post": "execute"})(
        _req(
            "post",
            _user(),
            {"pagination": {"host": {"page": 2, "page_size": 10}, "switch": {"page": 1, "page_size": 50}}},
        ),
        pk=7,
    )
    assert response.status_code == status.HTTP_200_OK
    assert captured["pagination"] == {"host": (2, 10), "switch": (1, 50)}


def test_execute_forwards_per_model_searches(monkeypatch):
    captured = {}
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: _scene())
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.execute_scene_query",
        lambda **kwargs: captured.update(kwargs) or {"total": 0, "models": []},
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.CmdbRulesFormatUtil.format_user_groups_permissions",
        lambda request, model_id: {},
    )
    response = SceneViewViewSet.as_view({"post": "execute"})(
        _req(
            "post",
            _user(),
            {
                "searches": {
                    "host": "  10.11  ",
                    "switch": "",
                    " ": "x",
                    "router": "a" * 200,
                }
            },
        ),
        pk=7,
    )
    assert response.status_code == status.HTTP_200_OK
    assert captured["searches"]["host"] == "10.11"
    assert "switch" not in captured["searches"]
    assert " " not in captured["searches"]
    assert len(captured["searches"]["router"]) == 128


def test_execute_forwards_typed_attr_searches(monkeypatch):
    captured = {}
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: _scene())
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.execute_scene_query",
        lambda **kwargs: captured.update(kwargs) or {"total": 0, "models": []},
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.CmdbRulesFormatUtil.format_user_groups_permissions",
        lambda request, model_id: {},
    )
    response = SceneViewViewSet.as_view({"post": "execute"})(
        _req(
            "post",
            _user(),
            {
                "searches": {
                    "host": {"field": "ip_addr", "type": "str*", "value": "10.11"},
                    "switch": ["bad"],
                }
            },
        ),
        pk=7,
    )
    assert response.status_code == status.HTTP_200_OK
    assert captured["searches"]["host"] == {"field": "ip_addr", "type": "str*", "value": "10.11"}
    assert "switch" not in captured["searches"]


def test_execute_rejects_invalid_searches(monkeypatch):
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: _scene())
    response = SceneViewViewSet.as_view({"post": "execute"})(
        _req("post", _user(), {"searches": ["host"]}),
        pk=7,
    )
    assert response.status_code == status.HTTP_400_BAD_REQUEST


def test_instance_viewer_can_create_personal_view(monkeypatch):
    monkeypatch.setattr(SceneViewSerializer, "save", lambda self, **kwargs: _scene())
    visitor = _user(
        is_superuser=False,
        permission={"cmdb": {"asset_info-View"}},
        roles=["cmdb_normal"],
    )
    response = SceneViewViewSet.as_view({"post": "create"})(_req("post", visitor, _payload()))
    assert response.status_code == status.HTTP_200_OK


def test_owner_with_view_permission_can_update_and_delete(monkeypatch):
    scene = _scene(created_by="bob")
    scene.delete = MagicMock()
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: scene)
    monkeypatch.setattr(SceneViewSerializer, "save", lambda self, **kwargs: scene)
    visitor = _user(
        username="bob",
        is_superuser=False,
        permission={"cmdb": {"asset_info-View"}},
        roles=["cmdb_normal"],
    )
    updated = SceneViewViewSet.as_view({"put": "update"})(_req("put", visitor, _payload()), pk=7)
    assert updated.status_code == status.HTTP_200_OK
    deleted = SceneViewViewSet.as_view({"delete": "destroy"})(_req("delete", visitor), pk=7)
    assert deleted.status_code == status.HTTP_200_OK
    scene.delete.assert_called_once()


def test_view_requires_asset_info_view():
    visitor = _user(is_superuser=False, permission={"cmdb": set()})
    response = SceneViewViewSet.as_view({"get": "list"})(_req("get", visitor))
    assert response.status_code == status.HTTP_403_FORBIDDEN


def test_list_includes_share_capabilities(monkeypatch):
    class _QS(list):
        def order_by(self, *_args):
            return self

    monkeypatch.setattr(SceneView.objects, "filter", lambda *args, **kwargs: _QS())
    admin = SceneViewViewSet.as_view({"get": "list"})(_req("get", _user()))
    assert _body(admin)["data"]["capabilities"] == {"can_org_share": True, "can_global": True}

    visitor = _user(is_superuser=False, permission={"cmdb": {"asset_info-View"}}, roles=["cmdb_normal"])
    listed = SceneViewViewSet.as_view({"get": "list"})(_req("get", visitor))
    assert _body(listed)["data"]["capabilities"] == {"can_org_share": False, "can_global": False}


def test_saved_log_uses_stable_template_and_omits_tags(monkeypatch, caplog):
    scene = _scene(tags=["env:secret"])
    monkeypatch.setattr(
        SceneViewSerializer,
        "save",
        lambda self, **kwargs: scene,
    )
    caplog.set_level(logging.INFO, logger="cmdb")
    created = SceneViewViewSet.as_view({"post": "create"})(_req("post", _user(), _payload(tags=["env:secret"])))
    records = [item for item in caplog.records if "event=scene_view_saved" in item.msg]
    assert len(records) == 1
    assert records[0].msg == "event=scene_view_saved scene_id=%s visibility=%s model_count=%s"
    assert records[0].args == (7, "personal", 2)
    assert "env:secret" not in caplog.text
    assert created.status_code == status.HTTP_200_OK
    assert _body(created)["result"] is True


def test_serializer_accepts_org_when_request_present():
    request = SimpleNamespace(user=_user())
    serializer = SceneViewSerializer(
        data=_payload(visibility="organization"),
        context={"request": request},
    )
    assert serializer.is_valid(), serializer.errors


def test_org_share_without_permission_is_403(monkeypatch):
    monkeypatch.setattr(SceneViewSerializer, "save", lambda self, **kwargs: _scene(visibility="organization"))
    visitor = _user(is_superuser=False, permission={"cmdb": {"asset_info-View"}}, roles=["cmdb_normal"])
    response = SceneViewViewSet.as_view({"post": "create"})(_req("post", visitor, _payload(visibility="organization"), cookies={"current_team": "1"}))
    assert response.status_code == status.HTTP_403_FORBIDDEN


def test_org_create_binds_current_team(monkeypatch):
    captured = {}

    def fake_save(self, **kwargs):
        captured.update(kwargs)
        return _scene(visibility="organization", organization=kwargs.get("organization"))

    monkeypatch.setattr(SceneViewSerializer, "save", fake_save)
    response = SceneViewViewSet.as_view({"post": "create"})(
        _req("post", _user(), _payload(visibility="organization"), cookies={"current_team": "12"})
    )
    assert response.status_code == status.HTTP_200_OK
    assert captured["visibility"] == "organization"
    assert captured["organization"] == 12


def test_global_without_admin_is_403(monkeypatch):
    monkeypatch.setattr(SceneViewSerializer, "save", lambda self, **kwargs: _scene(visibility="global"))
    visitor = _user(is_superuser=False, permission={"cmdb": {"asset_info-View"}}, roles=["cmdb_normal"])
    response = SceneViewViewSet.as_view({"post": "create"})(_req("post", visitor, _payload(visibility="global")))
    assert response.status_code == status.HTTP_403_FORBIDDEN


def test_non_owner_cannot_update_org_scene(monkeypatch):
    scene = _scene(visibility=SceneView.Visibility.ORGANIZATION, created_by="alice", organization=1)
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: scene)
    visitor = _user(
        username="bob",
        is_superuser=False,
        permission={"cmdb": {"asset_info-View", "asset_views_scene-Org Share"}},
        roles=["cmdb_normal"],
    )
    response = SceneViewViewSet.as_view({"put": "update"})(
        _req("put", visitor, _payload(visibility="organization"), cookies={"current_team": "1"}),
        pk=7,
    )
    assert response.status_code == status.HTTP_403_FORBIDDEN


def test_save_as_always_creates_personal_copy(monkeypatch):
    source = _scene(visibility=SceneView.Visibility.GLOBAL, created_by="alice", tags=["env:secret"])
    captured = {}
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: source)

    def fake_create(**kwargs):
        captured.update(kwargs)
        return _scene(id=99, **kwargs)

    monkeypatch.setattr(SceneView.objects, "create", fake_create)
    visitor = _user(username="bob", is_superuser=False, permission={"cmdb": {"asset_info-View"}}, roles=["cmdb_normal"])
    response = SceneViewViewSet.as_view({"post": "save_as"})(
        _req("post", visitor, {"name": "我的副本"}),
        pk=7,
    )
    assert response.status_code == status.HTTP_200_OK
    assert captured["visibility"] == SceneView.Visibility.PERSONAL
    assert captured["created_by"] == "bob"
    assert captured["organization"] is None
    assert captured["model_ids"] == ["host", "physcial_server"]
    assert captured["tags"] == ["env:secret"]
    assert _body(response)["data"]["visibility"] == "personal"


def test_export_builds_one_sheet_per_hit_model(monkeypatch):
    from io import BytesIO

    from openpyxl import Workbook, load_workbook

    scene = _scene()
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: scene)
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.collect_all_scene_instances",
        lambda **kwargs: {
            "total": 3,
            "models": [
                {"model_id": "host", "count": 2, "insts": [{"_id": 1}, {"_id": 2}]},
                {"model_id": "switch", "count": 1, "insts": [{"_id": 3}]},
            ],
        },
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view._model_table_columns",
        lambda model_id, creator: [{"attr_id": "inst_name", "attr_name": "名称", "attr_type": "str"}],
    )
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.CmdbRulesFormatUtil.format_user_groups_permissions",
        lambda request, model_id: {},
    )

    def fake_export(*, model_id, ids, **_kwargs):
        workbook = Workbook()
        workbook.active.append(["inst_name"])
        workbook.active.append([model_id])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    monkeypatch.setattr("apps.cmdb.views.scene_view.InstanceManage.inst_export", fake_export)
    response = SceneViewViewSet.as_view({"post": "export"})(_req("post", _user(), {}), pk=7)
    assert response.status_code == status.HTTP_200_OK
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["host", "switch"]


def test_export_empty_is_400(monkeypatch):
    monkeypatch.setattr(SceneViewViewSet, "get_object", lambda self: _scene())
    monkeypatch.setattr(
        "apps.cmdb.views.scene_view.collect_all_scene_instances",
        lambda **kwargs: {"total": 0, "models": []},
    )
    response = SceneViewViewSet.as_view({"post": "export"})(_req("post", _user(), {}), pk=7)
    assert response.status_code == status.HTTP_400_BAD_REQUEST


def test_export_unknown_scene_is_404(monkeypatch):
    from django.http import Http404

    monkeypatch.setattr(SceneViewViewSet, "get_object", MagicMock(side_effect=Http404))
    response = SceneViewViewSet.as_view({"post": "export"})(_req("post", _user(), {}), pk=404)
    assert response.status_code == status.HTTP_404_NOT_FOUND


def test_tag_options_unions_and_dedupes(monkeypatch):
    def fake_search(model_id, language="en"):
        if model_id == "host":
            return [
                {
                    "attr_id": "tag",
                    "attr_type": "tag",
                    "option": {
                        "options": [
                            {"key": "env", "value": "test"},
                            {"key": "app", "value": "shop"},
                        ]
                    },
                }
            ]
        return [
            {
                "attr_id": "tag",
                "attr_type": "tag",
                "option": {
                    "options": [
                        {"key": "env", "value": "test"},
                        {"key": "env", "value": "uat"},
                    ]
                },
            }
        ]

    monkeypatch.setattr("apps.cmdb.views.scene_view.ModelManage.search_model_attr", fake_search)
    response = SceneViewViewSet.as_view({"get": "tag_options"})(_req("get", _user(), query="?model_ids=host,switch"))
    assert response.status_code == status.HTTP_200_OK
    assert _body(response)["data"]["tags"] == ["env:test", "app:shop", "env:uat"]


def test_tag_options_only_uses_model_tag_field_not_string_or_instance_shape(monkeypatch):
    def fake_search(model_id, language="en"):
        if model_id == "qcloud_eip":
            return [
                {
                    "attr_id": "tag",
                    "attr_type": "str",
                    "option": {"validation_type": "unrestricted", "widget_type": "single_line"},
                }
            ]
        return [
            {
                "attr_id": "tag",
                "attr_type": "tag",
                "option": {"mode": "free", "options": [{"key": "test", "value": "aaa"}, {"key": "aaa", "value": "123"}]},
            }
        ]

    monkeypatch.setattr("apps.cmdb.views.scene_view.ModelManage.search_model_attr", fake_search)
    response = SceneViewViewSet.as_view({"get": "tag_options"})(_req("get", _user(), query="?model_ids=host,qcloud_eip"))
    assert response.status_code == status.HTTP_200_OK
    assert _body(response)["data"]["tags"] == ["test:aaa", "aaa:123"]


def test_menu_declares_org_share_and_normal_role_lacks_it():
    from pathlib import Path

    menu_path = Path(__file__).resolve().parents[3] / "support-files/system_mgmt/menus/cmdb.json"
    data = json.loads(menu_path.read_text())
    found = None
    for group in data["menus"]:
        for child in group.get("children") or []:
            if child.get("id") == "asset_views_scene":
                found = child
    assert found is not None
    assert "Org Share" in found["operation"]
    normal = next(role for role in data["roles"] if role["name"] == "normal")
    assert "asset_views_scene-Org Share" not in normal["menus"]
