"""场景视图活查询：AND/OR 标签条件、权限参数、省略 0 条模型、可见范围与导出。"""

from io import BytesIO

import pytest
from django.db.models import Q
from openpyxl import Workbook, load_workbook

from apps.cmdb.constants.field_constraints import TAG_ATTR_ID
from apps.cmdb.models.scene_view import SceneView
from apps.cmdb.services.scene_view import (
    build_inst_name_search,
    build_model_search_clause,
    build_tag_query_list,
    build_visible_scene_query,
    can_edit_scene,
    can_publish_global,
    can_publish_org,
    collect_all_scene_instances,
    execute_scene_query,
    merge_model_workbooks,
)

pytestmark = pytest.mark.unit


def test_and_emits_one_exact_tag_clause_per_value():
    clauses = build_tag_query_list(["env:test", "app:shop"], "and")
    assert clauses == [
        {"field": TAG_ATTR_ID, "type": "list_any[]", "value": ["env:test"], "accurate": True},
        {"field": TAG_ATTR_ID, "type": "list_any[]", "value": ["app:shop"], "accurate": True},
    ]


def test_or_emits_single_list_any_clause():
    clauses = build_tag_query_list(["env:test", "env:uat"], "or")
    assert clauses == [
        {
            "field": TAG_ATTR_ID,
            "type": "list_any[]",
            "value": ["env:test", "env:uat"],
            "accurate": True,
        }
    ]


def test_blank_tags_are_dropped_and_empty_means_no_tag_filter():
    assert build_tag_query_list(["", "  "], "and") == []
    assert build_tag_query_list([], "or") == []


def test_execute_omits_zero_count_models_and_sums_total():
    calls = []

    def fake_instance_list(*, model_id, params, page, page_size, order, permission_map, creator, **_kwargs):
        calls.append(
            {
                "model_id": model_id,
                "params": list(params),
                "permission_map": permission_map,
                "creator": creator,
                "page": page,
                "page_size": page_size,
                "order": order,
            }
        )
        if model_id == "host":
            return ([{"inst_uuid": "h1", "inst_name": "host-1", "model_id": "host"}], 2)
        return ([], 0)

    result = execute_scene_query(
        model_ids=["host", "switch"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        page=1,
        page_size=20,
        permission_map_loader=lambda model_id: {model_id: {"view"}},
        instance_list_fn=fake_instance_list,
    )

    assert result == {
        "total": 2,
        "models": [
            {
                "model_id": "host",
                "count": 2,
                "insts": [{"inst_uuid": "h1", "inst_name": "host-1", "model_id": "host"}],
            }
        ],
    }
    assert [item["model_id"] for item in calls] == ["host", "switch"]
    assert calls[0]["permission_map"] == {"host": {"view"}}
    assert calls[1]["permission_map"] == {"switch": {"view"}}
    assert calls[0]["creator"] == "alice"
    assert calls[0]["params"] == [{"field": TAG_ATTR_ID, "type": "list_any[]", "value": ["env:test"], "accurate": True}]
    assert calls[1]["params"] == calls[0]["params"]


def test_execute_uses_per_model_pagination():
    calls = []

    def fake_instance_list(*, model_id, params, page, page_size, **_kwargs):
        calls.append({"model_id": model_id, "page": page, "page_size": page_size})
        return ([{"inst_uuid": model_id}], 40)

    execute_scene_query(
        model_ids=["host", "switch"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        page=1,
        page_size=20,
        pagination={"host": (2, 10), "switch": (3, 50)},
        permission_map_loader=lambda _model_id: {},
        instance_list_fn=fake_instance_list,
    )

    assert calls == [
        {"model_id": "host", "page": 2, "page_size": 10},
        {"model_id": "switch", "page": 3, "page_size": 50},
    ]


def test_inst_name_search_is_contains_and_ignores_blank():
    assert build_inst_name_search("  10.11  ") == {
        "field": "inst_name",
        "type": "str*",
        "value": "10.11",
    }
    assert build_inst_name_search("   ") is None
    assert build_inst_name_search(None) is None
    clause = build_inst_name_search("a" * 200)
    assert clause is not None
    assert len(clause["value"]) == 128


def test_model_search_clause_accepts_attr_filter_and_legacy_keyword():
    assert build_model_search_clause("  10.11  ") == {
        "field": "inst_name",
        "type": "str*",
        "value": "10.11",
    }
    assert build_model_search_clause({"field": "ip_addr", "type": "str*", "value": "  10.11  "}) == {
        "field": "ip_addr",
        "type": "str*",
        "value": "10.11",
    }
    assert build_model_search_clause({"field": "owner", "type": "list[]", "value": ["alice", " bob ", ""]}) == {
        "field": "owner",
        "type": "list[]",
        "value": ["alice", "bob"],
    }
    assert build_model_search_clause({"field": "tag", "type": "list_any[]", "value": ["env:test"], "accurate": True}) == {
        "field": "tag",
        "type": "list_any[]",
        "value": ["env:test"],
        "accurate": True,
    }
    assert build_model_search_clause({"field": "n.hack", "type": "str*", "value": "x"}) is None
    assert build_model_search_clause({"field": "ip_addr", "type": "str<>", "value": "x"}) is None


def test_execute_appends_inst_name_search_per_model_and_keeps_zero_count():
    calls = []

    def fake_instance_list(*, model_id, params, **_kwargs):
        calls.append({"model_id": model_id, "params": list(params)})
        if model_id == "host":
            return ([], 0)
        return ([{"inst_uuid": "s1"}], 1)

    result = execute_scene_query(
        model_ids=["host", "switch"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        page=1,
        page_size=20,
        searches={"host": "  10.11  ", "switch": ""},
        permission_map_loader=lambda _model_id: {},
        instance_list_fn=fake_instance_list,
    )

    assert result["total"] == 1
    assert [item["model_id"] for item in result["models"]] == ["host", "switch"]
    assert result["models"][0]["count"] == 0
    assert calls[0]["params"] == [
        {"field": TAG_ATTR_ID, "type": "list_any[]", "value": ["env:test"], "accurate": True},
        {"field": "inst_name", "type": "str*", "value": "10.11"},
    ]
    assert calls[1]["params"] == [
        {"field": TAG_ATTR_ID, "type": "list_any[]", "value": ["env:test"], "accurate": True},
    ]


def test_execute_appends_typed_attr_search_clause():
    calls = []

    def fake_instance_list(*, model_id, params, case_sensitive=True, **_kwargs):
        calls.append({"params": list(params), "case_sensitive": case_sensitive})
        return ([{"inst_uuid": "h1"}], 1)

    execute_scene_query(
        model_ids=["host"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        searches={"host": {"field": "ip_addr", "type": "str*", "value": "10.11"}},
        permission_map_loader=lambda _model_id: {},
        instance_list_fn=fake_instance_list,
    )

    assert calls[0]["case_sensitive"] is False
    assert calls[0]["params"][-1] == {"field": "ip_addr", "type": "str*", "value": "10.11"}


def test_execute_copies_params_so_model_id_append_cannot_leak():
    seen = []

    def mutating_list(*, model_id, params, **_kwargs):
        params.append({"field": "model_id", "type": "str=", "value": model_id})
        seen.append(list(params))
        return ([{"inst_uuid": model_id}], 1)

    execute_scene_query(
        model_ids=["host", "server"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        page=1,
        page_size=10,
        permission_map_loader=lambda _model_id: {},
        instance_list_fn=mutating_list,
    )

    assert [item[-1]["value"] for item in seen] == ["host", "server"]
    assert all(item[0]["value"] == ["env:test"] for item in seen)


def _flatten_q(query: Q) -> list[tuple]:
    found: list[tuple] = []
    for child in query.children:
        if isinstance(child, Q):
            found.extend(_flatten_q(child))
        elif isinstance(child, tuple) and len(child) == 2:
            found.append(child)
    return found


def test_visible_query_isolates_org_and_keeps_global():
    query = build_visible_scene_query(username="alice", domain="domain.com", org_ids=[1])
    lookups = dict(_flatten_q(query))
    assert lookups["created_by"] == "alice"
    assert lookups["domain"] == "domain.com"
    assert lookups["organization__in"] == [1]
    assert {value for key, value in _flatten_q(query) if key == "visibility"} == {
        SceneView.Visibility.PERSONAL,
        SceneView.Visibility.ORGANIZATION,
        SceneView.Visibility.GLOBAL,
    }

    empty_orgs = build_visible_scene_query(username="alice", domain="domain.com", org_ids=[])
    empty_lookups = dict(_flatten_q(empty_orgs))
    assert empty_lookups.get("pk__in") == []
    assert "organization__in" not in empty_lookups


def test_org_share_and_global_gates():
    class User:
        def __init__(self, **kwargs):
            self.__dict__.update(kwargs)

    visitor = User(
        username="bob",
        is_superuser=False,
        permission={"cmdb": {"asset_info-View"}},
        roles=["cmdb_normal"],
    )
    sharer = User(
        username="cara",
        is_superuser=False,
        permission={"cmdb": {"asset_info-View", "asset_views_scene-Org Share"}},
        roles=["cmdb_normal"],
    )
    admin = User(username="ada", is_superuser=False, permission={"cmdb": set()}, roles=["admin"])
    assert can_publish_org(visitor) is False
    assert can_publish_org(sharer) is True
    assert can_publish_global(visitor) is False
    assert can_publish_global(admin) is True

    others_org = User(username="bob")
    others_org.username = "bob"
    scene = type("S", (), {"created_by": "alice", "visibility": SceneView.Visibility.ORGANIZATION})()
    global_scene = type("S", (), {"created_by": "alice", "visibility": SceneView.Visibility.GLOBAL})()
    assert can_edit_scene(visitor, scene) is False
    assert can_edit_scene(admin, global_scene) is True
    assert can_edit_scene(User(username="alice"), scene) is True


def test_collect_all_omits_zero_and_pages_through():
    calls = []

    def fake_instance_list(*, model_id, params, page, page_size, **_kwargs):
        calls.append({"model_id": model_id, "page": page, "page_size": page_size, "params": list(params)})
        if model_id == "empty":
            return ([], 0)
        if model_id == "host":
            if page == 1:
                return ([{"_id": 1}, {"_id": 2}], 3)
            return ([{"_id": 3}], 3)
        return ([{"_id": 9}], 1)

    result = collect_all_scene_instances(
        model_ids=["host", "empty", "switch"],
        tags=["env:test"],
        tag_match="and",
        creator="alice",
        permission_map_loader=lambda model_id: {model_id: True},
        instance_list_fn=fake_instance_list,
        page_size=2,
    )
    assert result["total"] == 4
    assert [item["model_id"] for item in result["models"]] == ["host", "switch"]
    assert [inst["_id"] for inst in result["models"][0]["insts"]] == [1, 2, 3]
    assert [item["page"] for item in calls if item["model_id"] == "host"] == [1, 2]


def test_merge_model_workbooks_keeps_hit_sheets_only():
    def sheet(rows):
        workbook = Workbook()
        active = workbook.active
        for row in rows:
            active.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    merged = merge_model_workbooks(
        [
            ("host", sheet([["name"], ["h1"]])),
            ("switch", sheet([["name"], ["s1"]])),
        ]
    )
    workbook = load_workbook(merged)
    assert workbook.sheetnames == ["host", "switch"]
    assert [cell.value for cell in workbook["host"][1]] == ["name"]
    assert [cell.value for cell in workbook["host"][2]] == ["h1"]
