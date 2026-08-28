from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.cmdb.services.application_resource_overview import ApplicationResourceOverviewService


@pytest.fixture(autouse=True)
def stub_model_name(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_info",
        lambda model_id: {"model_name": {"application": "应用", "host": "主机"}.get(model_id, model_id)},
    )


@pytest.mark.unit
def test_list_system_applications_filters_contains_applications(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.instance_association_instance_list",
        lambda model_id, inst_id: [
            {
                "src_model_id": "system",
                "dst_model_id": "application",
                "asst_id": "contains",
                "model_asst_id": "system_contains_application",
                "inst_list": [
                    {"_id": 11, "inst_uuid": "uuid-app-a", "model_id": "application", "inst_name": "app-a"},
                    {"_id": 12, "inst_uuid": "uuid-app-b", "model_id": "application", "inst_name": "app-b"},
                ],
            },
            {
                "src_model_id": "system",
                "dst_model_id": "host",
                "asst_id": "run",
                "model_asst_id": "system_run_host",
                "inst_list": [
                    {"_id": 21, "inst_uuid": "uuid-host-a", "model_id": "host", "inst_name": "host-a"},
                ],
            },
        ],
    )

    result = ApplicationResourceOverviewService.list_system_applications(100)

    assert [item["id"] for item in result] == ["uuid-app-a", "uuid-app-b"]
    assert [item["name"] for item in result] == ["app-a", "app-b"]


@pytest.mark.unit
def test_build_application_topology_returns_center_and_expanded_children(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_id",
        lambda inst_id: {
            "_id": inst_id,
            "inst_uuid": f"uuid-{inst_id}",
            "model_id": "application",
            "inst_name": "app-a",
        },
    )

    edges = {
        ("application", 11): [
            {
                "src_model_id": "application",
                "dst_model_id": "host",
                "asst_id": "run",
                "model_asst_id": "application_run_host",
                "inst_list": [{"_id": 21, "inst_uuid": "uuid-21", "model_id": "host", "inst_name": "host-a"}],
            },
            {
                "src_model_id": "application",
                "dst_model_id": "mysql",
                "asst_id": "contains",
                "model_asst_id": "application_contains_mysql",
                "inst_list": [{"_id": 31, "inst_uuid": "uuid-31", "model_id": "mysql", "inst_name": "db-a"}],
            },
        ],
        ("host", 21): [
            {
                "src_model_id": "host",
                "dst_model_id": "rack",
                "asst_id": "contains",
                "model_asst_id": "rack_contains_host",
                "inst_list": [{"_id": 41, "inst_uuid": "uuid-41", "model_id": "rack", "inst_name": "rack-a"}],
            },
        ],
        ("mysql", 31): [],
        ("rack", 41): [],
    }
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.instance_association_instance_list",
        lambda model_id, inst_id: edges.get((model_id, inst_id), []),
    )

    result = ApplicationResourceOverviewService.build_application_topology(
        inst_id=11,
        model_id="application",
        depth=2,
    )

    assert result["center"]["id"] == "uuid-11"
    assert {node["id"] for node in result["nodes"]} == {"uuid-11", "uuid-21", "uuid-31", "uuid-41"}
    assert result["links"]
    assert {link["source"] for link in result["links"]} <= {"uuid-11", "uuid-21", "uuid-31", "uuid-41"}
    assert {link["target"] for link in result["links"]} <= {"uuid-11", "uuid-21", "uuid-31", "uuid-41"}


@pytest.mark.unit
def test_build_application_resources_groups_by_category(monkeypatch):
    monkeypatch.setattr(
        ApplicationResourceOverviewService,
        "build_application_topology",
        staticmethod(
            lambda inst_id, model_id, depth=1, permission_map=None, user=None: {
                "center": {"id": "uuid-11", "model_id": "application", "name": "app-a"},
                "nodes": [
                    {"id": "uuid-11", "model_id": "application", "name": "app-a"},
                    {"id": "uuid-21", "model_id": "host", "name": "host-a"},
                    {"id": "uuid-31", "model_id": "mysql", "name": "db-a"},
                    {"id": "uuid-41", "model_id": "rack", "name": "rack-a"},
                ],
                "links": [],
                "truncated": False,
            }
        ),
    )

    result = ApplicationResourceOverviewService.build_application_resources(
        inst_id=11,
        model_id="application",
    )

    assert result["groups"]["application"][0]["id"] == "uuid-11"
    assert result["groups"]["host"][0]["id"] == "uuid-21"
    assert result["groups"]["database"][0]["id"] == "uuid-31"
    assert result["groups"]["rack_room"][0]["id"] == "uuid-41"


@pytest.mark.unit
def test_build_topology_instance_groups_returns_full_instances_grouped_by_model(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: (
            [
                {"attr_id": "inst_name", "attr_name": "实例名称", "attr_type": "str", "model_id": model_id},
                {"attr_id": "ip_addr", "attr_name": "IP地址", "attr_type": "str", "model_id": model_id},
            ]
            if model_id == "host"
            else [
                {"attr_id": "inst_name", "attr_name": "实例名称", "attr_type": "str", "model_id": model_id},
                {"attr_id": "owner", "attr_name": "负责人", "attr_type": "str", "model_id": model_id},
            ]
        ),
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [
            {"_id": 11, "model_id": "application", "inst_name": "app-a", "owner": "alice"},
            {"_id": 21, "model_id": "host", "inst_name": "host-a", "ip_addr": "10.0.0.1"},
            {"_id": 31, "model_id": "host", "inst_name": "host-b", "ip_addr": "10.0.0.2"},
        ],
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["21", "11", "31"])

    assert result["total"] == 3
    assert [group["model_id"] for group in result["groups"]] == ["application", "host"]
    assert result["groups"][0]["columns"] == ["inst_name", "owner"]
    assert result["groups"][1]["columns"] == ["inst_name", "ip_addr"]
    assert result["groups"][0]["column_defs"][0]["title"] == "实例名称"
    assert result["groups"][1]["column_defs"][-1]["title"] == "IP地址"
    assert result["groups"][0]["model_name"] == "应用"
    assert result["groups"][1]["items"][0]["inst_name"] == "host-a"
    assert "inst_uuid" not in result["groups"][1]["columns"]


@pytest.mark.unit
def test_export_topology_instance_groups_excel_creates_sheet_per_model(monkeypatch):
    monkeypatch.setattr(
        ApplicationResourceOverviewService,
        "build_topology_instance_groups",
        staticmethod(
            lambda node_ids, permission_map=None, user=None: {
                "groups": [
                    {
                        "model_id": "application",
                        "columns": ["inst_name", "model_id"],
                        "column_defs": [
                            {"key": "inst_name", "title": "实例名称"},
                            {"key": "model_id", "title": "模型ID"},
                        ],
                        "count": 1,
                        "items": [{"inst_name": "app-a", "model_id": "application"}],
                    },
                    {
                        "model_id": "host",
                        "columns": ["inst_name", "model_id"],
                        "column_defs": [
                            {"key": "inst_name", "title": "实例名称"},
                            {"key": "model_id", "title": "模型ID"},
                        ],
                        "count": 1,
                        "items": [{"inst_name": "host-a", "model_id": "host"}],
                    },
                ],
                "total": 2,
            }
        ),
    )

    content = ApplicationResourceOverviewService.export_topology_instance_groups_excel(["11", "21"])
    workbook = load_workbook(filename=BytesIO(content))

    assert workbook.sheetnames == ["application", "host"]
    assert workbook["application"]["A1"].value == "实例名称"
    assert workbook["application"]["A2"].value == "app-a"
    assert workbook["host"]["A2"].value == "host-a"


@pytest.mark.unit
def test_build_topology_instance_groups_prefers_display_values(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {
                "attr_id": "status",
                "attr_name": "状态",
                "attr_type": "enum",
                "model_id": model_id,
                "option": [{"id": "1", "name": "运行中"}],
            },
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [{"_id": 11, "model_id": "application", "inst_name": "app-a", "status": "1"}],
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["11"])

    assert result["groups"][0]["column_defs"][-1]["title"] == "状态"
    assert result["groups"][0]["items"][0]["status"] == "运行中"


@pytest.mark.unit
def test_build_topology_instance_groups_translates_stringified_enum_list(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {
                "attr_id": "app_type",
                "attr_name": "应用类型",
                "attr_type": "enum",
                "model_id": model_id,
                "option": [
                    {"id": "1", "name": "基础应用"},
                    {"id": "2", "name": "业务应用"},
                ],
            },
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [{"_id": 11, "model_id": "application", "inst_name": "app-a", "app_type": "['2']"}],
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["11"])

    assert result["groups"][0]["column_defs"][-1]["title"] == "应用类型"
    assert result["groups"][0]["items"][0]["app_type"] == "业务应用"


@pytest.mark.unit
def test_build_topology_instance_groups_hides_display_fields(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {"attr_id": "inst_name", "attr_name": "实例名称", "attr_type": "str", "model_id": model_id},
            {"attr_id": "app_type", "attr_name": "应用类型", "attr_type": "enum", "model_id": model_id, "option": [{"id": "2", "name": "业务应用"}]},
            {"attr_id": "app_type_display", "attr_name": "应用类型", "attr_type": "str", "model_id": model_id, "is_display_field": True},
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [
            {
                "_id": 11,
                "model_id": "application",
                "inst_name": "app-a",
                "app_type": "['2']",
                "app_type_display": "业务应用",
            }
        ],
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["11"])

    assert result["groups"][0]["columns"] == ["inst_name", "app_type"]
    assert [item["title"] for item in result["groups"][0]["column_defs"]] == ["实例名称", "应用类型"]
    assert "app_type_display" not in result["groups"][0]["items"][0]
    assert result["groups"][0]["items"][0]["app_type"] == "业务应用"


@pytest.mark.unit
def test_build_topology_instance_groups_rebuilds_stale_display_fields(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {"attr_id": "app_type", "attr_name": "应用类型", "attr_type": "enum", "model_id": model_id, "option": [{"id": "2", "name": "数据库"}]},
            {"attr_id": "app_type_display", "attr_name": "应用类型", "attr_type": "str", "model_id": model_id, "is_display_field": True},
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [
            {
                "_id": 11,
                "model_id": "application",
                "inst_name": "app-a",
                "app_type": ["2"],
                "app_type_display": "['2']",
            }
        ],
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["11"])

    assert result["groups"][0]["items"][0]["app_type"] == "数据库"


@pytest.mark.unit
def test_build_topology_instance_groups_translates_cloud_region(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {"attr_id": "cloud", "attr_name": "云区域", "attr_type": "int", "model_id": model_id},
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [{"_id": 21, "model_id": "host", "inst_name": "host-a", "cloud": 2}],
    )

    class _FakeNodeMgmt:
        def cloud_region_list(self):
            return [{"id": 2, "name": "华南广州"}]

    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.NodeMgmt",
        lambda: _FakeNodeMgmt(),
    )
    ApplicationResourceOverviewService._cloud_region_name_cache = None

    result = ApplicationResourceOverviewService.build_topology_instance_groups(["21"])

    assert result["groups"][0]["column_defs"][-1]["title"] == "云区域"
    assert result["groups"][0]["items"][0]["cloud"] == "华南广州"


class _ShowFieldUser:
    username = "alice"


@pytest.mark.unit
def test_build_topology_instance_groups_uses_show_fields(monkeypatch):
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.ModelManage.search_model_attr_v2",
        lambda model_id: [
            {"attr_id": "inst_name", "attr_name": "实例名称", "attr_type": "str", "model_id": model_id},
            {"attr_id": "owner", "attr_name": "负责人", "attr_type": "user", "model_id": model_id},
            {"attr_id": "remark", "attr_name": "备注", "attr_type": "str", "model_id": model_id},
            {"attr_id": "cpu_model", "attr_name": "CPU型号", "attr_type": "str", "model_id": model_id},
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.query_entity_by_ids",
        lambda ids: [
            {
                "_id": 21,
                "model_id": "host",
                "inst_uuid": "uuid-host-a",
                "inst_name": "host-a",
                "owner": ["张备份"],
                "remark": "prod",
                "cpu_model": "Xeon",
                "created_at": "2026-06-16T08:06:22.629308",
            }
        ],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.application_resource_overview.InstanceManage.get_info",
        lambda model_id, created_by: {"show_fields": ["inst_name", "owner", "remark"]},
    )

    result = ApplicationResourceOverviewService.build_topology_instance_groups(
        ["21"],
        user=_ShowFieldUser(),
    )

    group = result["groups"][0]
    assert group["columns"] == ["inst_name", "owner", "remark"]
    assert "cpu_model" not in group["columns"]
    assert "created_at" not in group["columns"]
    assert group["items"][0]["owner"] == "张备份"
    assert group["items"][0]["inst_uuid"] == "uuid-host-a"
    assert group["items"][0]["model_id"] == "host"


@pytest.mark.unit
def test_serialize_instance_value_joins_lists_and_formats_time():
    assert ApplicationResourceOverviewService._serialize_instance_value(["张备份", "李运维"]) == "张备份, 李运维"
    assert ApplicationResourceOverviewService._serialize_instance_value("['张备份']") == "张备份"
    assert ApplicationResourceOverviewService._serialize_instance_value("2026-06-16T08:06:22.629308") == "2026-06-16 08:06:22"
    assert ApplicationResourceOverviewService._serialize_instance_value(None) == ""
