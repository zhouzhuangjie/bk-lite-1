import json
import os
import uuid
from dataclasses import asdict
from hashlib import sha256
from typing import Any, Callable

from django.core.cache import cache
from django.utils import timezone

from apps.cmdb.constants.constants import (
    CLASSIFICATION,
    CREATE_MODEL_CHECK_ATTR,
    DISPLAY_FIELD_CONFIG,
    ENUM_SELECT_MODE_DEFAULT,
    INST_NAME_INFOS,
    INSTANCE,
    MODEL,
    MODEL_ASSOCIATION,
    OPERATOR_MODEL,
    ORGANIZATION,
    SUBORDINATE_MODEL,
    UPDATE_MODEL_CHECK_ATTR_MAP,
    USER,
)
from apps.cmdb.constants.field_constraints import TAG_ATTR_ID, TAG_MODE_FREE
from apps.cmdb.custom_reporting.extensions import get_custom_reporting_extension
from apps.cmdb.display_field.constants import DISPLAY_FIELD_TYPES, DISPLAY_SUFFIX
from apps.cmdb.graph.drivers.graph_client import GraphClient
from apps.cmdb.graph.validators import MAX_BATCH_UPDATE_PROPERTY_VALUES
from apps.cmdb.language.service import SettingLanguage
from apps.cmdb.model_ops.extensions import get_model_enterprise_extension, is_file_attr_type
from apps.cmdb.models import CREATE_INST, DELETE_INST, UPDATE_INST, FieldGroup
from apps.cmdb.models.change_record import MODEL_MANAGEMENT_CHANGE
from apps.cmdb.services.auto_relation_rule import (
    AUTO_RELATION_RULE_FIELD,
    AutoRelationRule,
    AutoRelationRuleSet,
    build_auto_relation_rule_response,
    canonicalize_auto_relation_rule_set_payload,
    dump_auto_relation_rule_set,
    dump_auto_relation_rule_set_compact,
    parse_auto_relation_rule_set,
    validate_auto_relation_rule_payload,
    validate_auto_relation_rule_set_payload,
)
from apps.cmdb.services.classification import ClassificationManage
from apps.cmdb.services.model_attribute_policy import ModelAttributePolicy
from apps.cmdb.services.model_visibility import BusinessModelVisibility
from apps.cmdb.services.unique_rule import (
    UniqueRulePayload,
    apply_unique_rules_to_attr_export_rows,
    build_unique_rule_context,
    build_unique_rules_from_attr_rows,
    copy_unique_rules_to_model,
    create_unique_rule,
    delete_unique_rule,
    enrich_attrs_with_unique_display,
    guard_attr_change_against_unique_rules,
    list_unique_rule_candidate_fields,
    list_unique_rules,
    update_unique_rule,
    validate_unique_rules_against_existing_instances,
)
from apps.cmdb.utils.change_record import create_change_record
from apps.cmdb.validators import IdentifierValidator
from apps.core.exceptions.base_app_exception import BaseAppException
from apps.core.logger import cmdb_logger as logger
from apps.core.services.user_group import UserGroup
from apps.rpc.system_mgmt import SystemMgmt

FIELD_GROUP_MANAGER: Any = getattr(FieldGroup, "objects")
PROTECTED_MODEL_ATTR_IDS = frozenset({ORGANIZATION})
MODEL_ATTR_OPTION_CACHE_TTL = int(os.getenv("CMDB_MODEL_ATTR_OPTION_CACHE_TTL", "300"))
MODEL_ATTR_ORGANIZATION_OPTION_CACHE_KEY = "cmdb:model_attr_options:organization"
MODEL_ATTR_USER_OPTION_CACHE_KEY = "cmdb:model_attr_options:user"
ENUM_DISPLAY_BACKFILL_CHECKPOINT_TTL = 24 * 60 * 60


class ModelManage(object):
    @staticmethod
    def _enum_display_backfill_checkpoint(model_id: str, attr_id: str, new_options: list) -> tuple[str, str]:
        options_digest = sha256(json.dumps(new_options, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
        return f"cmdb:enum-display-backfill:{model_id}:{attr_id}", options_digest

    @staticmethod
    def _get_enum_display_backfill_checkpoint(checkpoint_key: str, options_digest: str) -> tuple[int | None, bool]:
        try:
            checkpoint = cache.get(checkpoint_key)
            if not isinstance(checkpoint, dict):
                return None, True
            if checkpoint.get("options_digest") != options_digest:
                cache.delete(checkpoint_key)
                return None, True
            return int(checkpoint["cursor"]), True
        except Exception as e:
            logger.warning(
                "[update_enum_instances_display] 读取回填断点失败，将从头幂等执行: error_type=%s, error=%s",
                type(e).__name__,
                e,
            )
            return None, False

    @staticmethod
    def _save_enum_display_backfill_checkpoint(checkpoint_key: str, options_digest: str, cursor: int) -> bool:
        try:
            cache.set(
                checkpoint_key,
                {"options_digest": options_digest, "cursor": cursor},
                timeout=ENUM_DISPLAY_BACKFILL_CHECKPOINT_TTL,
            )
            return True
        except Exception as e:
            logger.warning(
                "[update_enum_instances_display] 保存回填断点失败，后续重试将从头幂等执行: cursor=%s, error_type=%s, error=%s",
                cursor,
                type(e).__name__,
                e,
            )
            return False

    @staticmethod
    def _clear_enum_display_backfill_checkpoint(checkpoint_key: str) -> None:
        try:
            cache.delete(checkpoint_key)
        except Exception as e:
            logger.warning(
                "[update_enum_instances_display] 清理回填断点失败，后续会幂等续扫: error_type=%s, error=%s",
                type(e).__name__,
                e,
            )

    @staticmethod
    def _write_enum_display_batch(ag, display_field_id: str, property_values: list[dict]) -> int:
        for attempt in range(2):
            try:
                written = ag.batch_update_node_property_values(INSTANCE, display_field_id, property_values)
                return len(written)
            except Exception as e:
                if attempt:
                    raise
                logger.warning(
                    "[update_enum_instances_display] 图写失败，将从当前成功游标重试一次: error_type=%s, error=%s",
                    type(e).__name__,
                    e,
                )
        return 0

    @staticmethod
    def _clone_options(option: list[dict]) -> list[dict]:
        return [dict(item) for item in option]

    @staticmethod
    def _normalize_default_value(raw_value: Any) -> list[str]:
        return ModelAttributePolicy._normalize_default_value(raw_value)

    @staticmethod
    def sanitize_attr_default_value(attr: dict, log_context: str = "") -> dict:
        return ModelAttributePolicy._sanitize_attr_default_value(
            attr,
            log_context,
            ModelManage._normalize_default_value,
            ModelManage.resolve_runtime_enum_options,
            logger,
        )

    @staticmethod
    def normalize_enum_public_binding(attr_info: dict, current_attr: dict | None = None) -> dict:
        return ModelAttributePolicy._normalize_enum_public_binding(attr_info, current_attr, logger)

    @staticmethod
    def validate_enum_rule_immutable(current_attr: dict, incoming_attr: dict) -> None:
        return ModelAttributePolicy.validate_enum_rule_immutable(current_attr, incoming_attr)

    @staticmethod
    def ensure_enum_select_mode(attr_info: dict) -> dict:
        return ModelAttributePolicy.ensure_enum_select_mode(attr_info)

    @staticmethod
    def validate_enum_select_mode_immutable(current_attr: dict, incoming_attr: dict) -> None:
        return ModelAttributePolicy.validate_enum_select_mode_immutable(current_attr, incoming_attr)

    @staticmethod
    def resolve_runtime_enum_options(attr: dict) -> list[dict]:
        return ModelAttributePolicy._resolve_runtime_enum_options(attr, logger)

    @staticmethod
    def _normalize_attr_constraints(attrs: list[dict]) -> list[dict]:
        normalized: list[dict] = []
        for attr in attrs:
            item = dict(attr)
            item.setdefault("is_only", False)
            item.setdefault("is_required", False)
            item.setdefault("editable", True)
            item.setdefault("option", {})
            item.setdefault("user_prompt", "")
            item.setdefault("default_value", [])
            item = ModelManage.sanitize_attr_default_value(item, log_context="normalize_attr_constraints")
            normalized.append(item)
        return normalized

    @staticmethod
    def _is_tag_attr(attr: dict) -> bool:
        return ModelAttributePolicy._is_tag_attr(attr)

    @staticmethod
    def validate_tag_attr_definition(attrs: list[dict], incoming_attr: dict) -> None:
        return ModelAttributePolicy.validate_tag_attr_definition(attrs, incoming_attr)

    @staticmethod
    def normalize_tag_field_option(option: dict | list[Any] | None) -> dict:
        return ModelAttributePolicy.normalize_tag_field_option(option)

    @staticmethod
    def merge_tag_options_from_values(model_id: str, values: list[str]) -> None:
        if not values:
            return
        model_info = ModelManage.search_model_info(model_id)
        if not model_info:
            return
        attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
        tag_attr = next((attr for attr in attrs if ModelManage._is_tag_attr(attr)), None)
        if not tag_attr:
            return

        option = ModelManage.normalize_tag_field_option(tag_attr.get("option") or {})
        if option.get("mode") != TAG_MODE_FREE:
            return

        existing = {
            (str(item.get("key", "")).strip(), str(item.get("value", "")).strip()) for item in option.get("options", []) if isinstance(item, dict)
        }
        changed = False
        for raw in values:
            if not isinstance(raw, str) or raw.count(":") != 1:
                continue
            key, value = [part.strip() for part in raw.split(":", 1)]
            if not key or not value:
                continue
            pair = (key, value)
            if pair in existing:
                continue
            existing.add(pair)
            option["options"].append({"key": key, "value": value})
            changed = True

        if not changed:
            return

        for attr in attrs:
            if ModelManage._is_tag_attr(attr):
                attr["option"] = option
                break

        with GraphClient() as ag:
            ag.set_entity_properties(
                MODEL,
                [model_info["_id"]],
                {"attrs": json.dumps(attrs, ensure_ascii=False)},
                {},
                [],
                False,
            )

    @staticmethod
    def _validate_attr_id(attr_id: str):
        if not IdentifierValidator.is_valid(attr_id):
            raise BaseAppException(IdentifierValidator.get_error_message("属性ID"))

    @staticmethod
    def _validate_model_id(model_id: str):
        if not IdentifierValidator.is_valid(model_id):
            raise BaseAppException(IdentifierValidator.get_error_message("模型ID"))

    @staticmethod
    def _guard_protected_model_attr(attr_id: str, action: str):
        if attr_id in PROTECTED_MODEL_ATTR_IDS:
            raise BaseAppException(f"模型字段 {attr_id} 为系统字段，不允许{action}")

    @staticmethod
    def _add_display_field_to_attrs(attrs: list, attr_info: dict, model_id: str, is_pre: bool = False):
        """
        为指定属性添加 _display 字段定义
        Args:
            attrs: 属性列表
            attr_info: 属性信息字典
            model_id: 模型ID
            is_pre: 是否为预定义字段
        Returns:
            bool: 是否添加了 _display 字段
        """
        from apps.cmdb.display_field import DISPLAY_FIELD_TYPES, DISPLAY_SUFFIX

        attr_type = attr_info.get("attr_type")
        if attr_type not in DISPLAY_FIELD_TYPES:
            return False

        attr_id = attr_info["attr_id"]
        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"

        # 检查是否已存在 _display 字段
        if display_field_id in {a["attr_id"] for a in attrs}:
            return False

        # 创建 _display 字段定义
        display_field = {
            "attr_id": display_field_id,
            "attr_name": attr_info.get("attr_name"),
            "attr_group": attr_info.get("attr_group", "default"),
            "group_id": attr_info.get("group_id"),
            "model_id": model_id,
            **DISPLAY_FIELD_CONFIG,
        }
        if is_pre:
            display_field["is_pre"] = True

        attrs.append(display_field)
        return True

    @staticmethod
    def _remove_display_field_from_attrs(attrs: list, attr_id: str):
        """
        从属性列表中删除指定属性的 _display 字段定义
        Args:
            attrs: 属性列表
            attr_id: 原始属性ID
        Returns:
            tuple: (新的属性列表, 是否删除了 _display 字段)
        """
        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"
        original_count = len(attrs)
        new_attrs = [a for a in attrs if a["attr_id"] != display_field_id]

        return new_attrs, len(new_attrs) < original_count

    @staticmethod
    def _handle_attr_type_change(
        attrs: list,
        attr_id: str,
        old_type: str,
        new_type: str,
        attr_info: dict,
        model_id: str,
        graph_client,
    ):
        """
        处理属性类型变更时的 _display 字段逻辑
        Args:
            attrs: 属性列表
            attr_id: 属性ID
            old_type: 原类型
            new_type: 新类型
            attr_info: 属性信息
            model_id: 模型ID
            graph_client: 图数据库客户端实例
        Returns:
            list: 更新后的属性列表
        """
        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"

        # 情况1: 从非目标类型改为目标类型 -> 添加 _display 字段
        if old_type not in DISPLAY_FIELD_TYPES and new_type in DISPLAY_FIELD_TYPES:
            ModelManage._add_display_field_to_attrs(attrs, attr_info, model_id)

        # 情况2: 从目标类型改为非目标类型 -> 删除 _display 字段和实例数据
        elif old_type in DISPLAY_FIELD_TYPES and new_type not in DISPLAY_FIELD_TYPES:
            # 从 attrs 中删除 _display 字段定义
            attrs, removed = ModelManage._remove_display_field_from_attrs(attrs, attr_id)

            # 删除所有实例的 _display 字段数据
            if removed:
                model_params = [{"field": "model_id", "type": "str=", "value": model_id}]
                graph_client.remove_entitys_properties(INSTANCE, model_params, [display_field_id])

        return attrs

    @staticmethod
    def create_model(data: dict, username="admin"):
        """
        创建模型
        """

        attrs = list(INST_NAME_INFOS)  # 复制默认字段列表
        model_id = str(data.get("model_id", ""))

        # 为默认字段中的目标类型添加 _display 字段定义
        for attr in list(attrs):  # 使用 list() 避免迭代时修改列表
            ModelManage._add_display_field_to_attrs(attrs, attr, model_id, is_pre=True)

        data.update(attrs=json.dumps(attrs), unique_rules="[]")

        with GraphClient() as ag:
            # 仅查询与新模型存在唯一性冲突可能的节点（model_id 或 model_name 匹配），
            # 避免全量加载所有 MODEL 节点做内存比对（O(N) → O(冲突候选数)）。
            conflict_filter = [
                {"field": "model_id", "type": "str=", "value": data.get("model_id", "")},
                {"field": "model_name", "type": "str=", "value": data.get("model_name", "")},
            ]
            exist_items, _ = ag.query_entity(MODEL, conflict_filter, param_type="OR")
            result = ag.create_entity(MODEL, data, CREATE_MODEL_CHECK_ATTR, exist_items)
            classification_info = ClassificationManage.search_model_classification_info(data["classification_id"])
            _ = ag.create_edge(
                SUBORDINATE_MODEL,
                classification_info["_id"],
                CLASSIFICATION,
                result["_id"],
                MODEL,
                dict(
                    classification_model_asst_id=f"{result['classification_id']}_{SUBORDINATE_MODEL}_{result['model_id']}"
                    # noqa
                ),
                "classification_model_asst_id",
            )

        # 初始化排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(data["model_id"])

        # 为新建模型创建默认字段分组
        FIELD_GROUP_MANAGER.create(
            model_id=data["model_id"],
            group_name="default",
            order=1,
            is_collapsed=False,
            description="默认分组",
            created_by=username,
            attr_orders=[attr.get("attr_id") for attr in attrs if not attr.get("is_display_field")],
        )

        create_change_record(
            operator=username,
            model_id=data["model_id"],
            label="模型管理",
            _type=CREATE_INST,
            message=f"创建模型. 模型名称: {data['model_name']}",
            inst_id=result["_id"],
            model_object=OPERATOR_MODEL,
            scenario=MODEL_MANAGEMENT_CHANGE,
        )
        return result

    @staticmethod
    def register_custom_reporting_model_fields(
        model_id: str,
        instances: list[dict],
        username="admin",
    ) -> list[str]:
        return get_custom_reporting_extension().register_model_fields(
            model_id,
            instances,
            username=username,
        )

    @staticmethod
    def validate_custom_reporting_instance_fields(model_id: str, instances: list[dict]) -> None:
        get_custom_reporting_extension().validate_instance_fields(model_id, instances)

    @staticmethod
    def _get_custom_reporting_declared_attr_ids(model_id: str) -> set[str]:
        return get_custom_reporting_extension().get_declared_attr_ids(model_id)

    @staticmethod
    def validate_custom_reporting_relation_fields(
        model_id: str,
        relations: list[dict],
        identity_keys: list[str] | None = None,
    ) -> None:
        get_custom_reporting_extension().validate_relation_fields(
            model_id,
            relations,
            identity_keys=identity_keys,
        )

    @staticmethod
    def normalize_custom_reporting_identity_keys(identity_keys) -> list[str]:
        return get_custom_reporting_extension().normalize_identity_keys(identity_keys)

    @staticmethod
    def bootstrap_custom_reporting_model(quick_model: dict, team: list[int], username="admin"):
        return get_custom_reporting_extension().bootstrap_model(
            quick_model,
            team=team,
            username=username,
        )

    @staticmethod
    def sync_custom_reporting_model_group(quick_model: dict, team: list[int], username="admin"):
        return get_custom_reporting_extension().sync_model_group(
            quick_model,
            team=team,
            username=username,
        )

    @staticmethod
    def copy_model(
        src_model_id: str,
        new_model_id: str,
        new_model_name: str,
        classification_id: str | None = None,
        group: list | None = None,
        icn: str | None = None,
        copy_attributes: bool = False,
        copy_relationships: bool = False,
        username="admin",
    ):
        """
        复制模型
        Args:
            src_model_id: 源模型ID
            new_model_id: 新模型ID
            new_model_name: 新模型名称
            classification_id: 模型分类ID（可选，不传则继承源模型）
            group: 组织列表（可选，不传则继承源模型）
            icn: 图标（可选，不传则继承源模型）
            copy_attributes: 是否复制属性
            copy_relationships: 是否复制关系
            username: 操作用户
        Returns:
            新模型信息
        """
        # 校验新模型ID格式
        ModelManage._validate_model_id(new_model_id)

        # 校验复制方式
        if not copy_attributes and not copy_relationships:
            raise BaseAppException("至少选择一种复制方式（属性或关系）")

        # 获取源模型信息
        src_model_info = ModelManage.search_model_info(src_model_id)
        if not src_model_info:
            raise BaseAppException("源模型不存在")

        # 准备属性列表
        if copy_attributes:
            # 完全复制源模型的所有属性
            src_attrs = ModelManage.parse_attrs(src_model_info.get("attrs", "[]"))
            attrs = [dict(attr) for attr in src_attrs]
        else:
            # 不复制属性：只使用默认属性
            attrs = list(INST_NAME_INFOS)
            # 为默认字段中的目标类型添加 _display 字段定义
            for attr in list(attrs):
                ModelManage._add_display_field_to_attrs(attrs, attr, new_model_id, is_pre=True)

        # 构建新模型数据
        new_model_data = {
            "model_id": new_model_id,
            "model_name": new_model_name,
            "classification_id": classification_id or src_model_info.get("classification_id"),
            "group": group if group is not None else src_model_info.get("group", []),
            "icn": icn if icn is not None else src_model_info.get("icn", ""),
            "attrs": json.dumps(attrs),
            "unique_rules": "[]",
        }

        # 一次性创建模型（包含所有属性）
        with GraphClient() as ag:
            # 仅查询与新模型存在唯一性冲突可能的节点（model_id 或 model_name 匹配），
            # 避免全量加载所有 MODEL 节点做内存比对（O(N) → O(冲突候选数)）。
            conflict_filter = [
                {"field": "model_id", "type": "str=", "value": new_model_data.get("model_id", "")},
                {"field": "model_name", "type": "str=", "value": new_model_data.get("model_name", "")},
            ]
            exist_items, _ = ag.query_entity(MODEL, conflict_filter, param_type="OR")
            new_model = ag.create_entity(MODEL, new_model_data, CREATE_MODEL_CHECK_ATTR, exist_items)

            # 创建模型与分类的关联
            classification_info = ClassificationManage.search_model_classification_info(new_model_data["classification_id"])
            ag.create_edge(
                SUBORDINATE_MODEL,
                classification_info["_id"],
                CLASSIFICATION,
                new_model["_id"],
                MODEL,
                dict(classification_model_asst_id=f"{new_model['classification_id']}_{SUBORDINATE_MODEL}_{new_model['model_id']}"),
                "classification_model_asst_id",
            )

        # 初始化排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(new_model_id)

        try:
            # 处理字段分组复制
            if copy_attributes:
                # 复制源模型的字段分组配置
                src_field_groups = FIELD_GROUP_MANAGER.filter(model_id=src_model_id).order_by("order")
                for src_group in src_field_groups:
                    FIELD_GROUP_MANAGER.create(
                        model_id=new_model_id,
                        group_name=src_group.group_name,
                        attr_orders=src_group.attr_orders or [],
                        order=src_group.order,
                    )
                copy_unique_rules_to_model(src_model_id, new_model_id, username)
            else:
                # 不复制属性时，为新模型创建默认分组
                FIELD_GROUP_MANAGER.create(
                    model_id=new_model_id,
                    group_name="default",
                    order=1,
                    is_collapsed=False,
                    description="默认分组",
                    created_by=username,
                    attr_orders=[attr.get("attr_id") for attr in attrs if not attr.get("is_display_field")],
                )

            # 处理关系复制
            if copy_relationships:
                associations = ModelManage.model_association_search(src_model_id)

                for assoc in associations:
                    # 确定新的源模型ID和目标模型ID
                    new_src_model_id = new_model_id if assoc["src_model_id"] == src_model_id else assoc["src_model_id"]
                    new_dst_model_id = new_model_id if assoc["dst_model_id"] == src_model_id else assoc["dst_model_id"]

                    # 如果关联的两端都是源模型（自关联），则两端都改为新模型
                    if assoc["src_model_id"] == src_model_id and assoc["dst_model_id"] == src_model_id:
                        new_src_model_id = new_model_id
                        new_dst_model_id = new_model_id

                    # 获取新的src和dst的_id
                    src_model_info_new = ModelManage.search_model_info(new_src_model_id)
                    dst_model_info_new = ModelManage.search_model_info(new_dst_model_id)

                    if not src_model_info_new or not dst_model_info_new:
                        continue

                    # 创建新的关联ID
                    new_model_asst_id = f"{new_src_model_id}_{assoc['asst_id']}_{new_dst_model_id}"

                    # 复制关联关系
                    try:
                        ModelManage.model_association_create(
                            src_id=src_model_info_new["_id"],
                            dst_id=dst_model_info_new["_id"],
                            src_model_id=new_src_model_id,
                            dst_model_id=new_dst_model_id,
                            asst_id=assoc.get("asst_id", ""),
                            asst_name=assoc.get("asst_name", ""),
                            mapping=assoc.get("mapping", "1:n"),
                            on_delete=assoc.get("on_delete", "none"),
                            is_pre=assoc.get("is_pre", False),
                            model_asst_id=new_model_asst_id,
                        )
                    except BaseAppException:
                        # 如果关联已存在，跳过
                        continue

            # 记录变更
            create_change_record(
                operator=username,
                model_id=new_model_id,
                label="模型管理",
                _type=CREATE_INST,
                message=f"复制模型. 源模型: {src_model_info['model_name']}, 新模型: {new_model_name}",
                inst_id=new_model["_id"],
                model_object=OPERATOR_MODEL,
                scenario=MODEL_MANAGEMENT_CHANGE,
            )

            return new_model

        except Exception as e:
            # 如果复制过程中出错，删除已创建的模型
            try:
                ModelManage.delete_model(new_model["_id"])
            except Exception:  # noqa: BLE001 - 清理失败不应掩盖原始错误
                pass
            raise e

    @staticmethod
    def delete_model(id: int):
        """
        删除模型
        """
        with GraphClient() as ag:
            ag.batch_delete_entity(MODEL, [id])

    @staticmethod
    def update_model(id: int, data: dict):
        """
        更新模型
        TODO 不能单独更新一个字段，如只更新icon，传递全部字段会导致其他字段校验不通过 model_name 后续考虑优化
        """
        model_id = data.pop("model_id", "")  # 不能更新model_id
        with GraphClient() as ag:
            exist_items, _ = ag.query_entity(MODEL, [{"field": "model_id", "type": "str<>", "value": model_id}])
            # 排除当前正在更新的模型，避免自己和自己比较
            exist_items = [i for i in exist_items if i["_id"] != id]
            model = ag.set_entity_properties(MODEL, [id], data, UPDATE_MODEL_CHECK_ATTR_MAP, exist_items)
        return model[0]

    @staticmethod
    def search_model(
        language: str = "en",
        order_type: str = "ASC",
        order: str = "order_id",
        permissions_map: dict | None = None,
        classification_ids: list | None = None,
        creator: str = "",
        include_hidden: bool = False,
    ):
        """
        查询模型
        Args:
            language: 语言，默认英语
            order_type: 排序方式，asc升序/desc降序
            order: 排序字段，默认 order_id
            classification_ids: 分类ID列表，可选，用于过滤特定分类下的模型
            permissions_map: 权限过滤字典
            creator: 创建人，可选
            include_hidden: True 时返回包含已隐藏（is_visible=False）的模型；
                默认 False 时过滤掉已隐藏项
        """

        permissions_map = permissions_map or {}
        format_permission_dict = {}

        for organization_id, organization_permission_data in permissions_map.items():
            _query_list = []
            if classification_ids:
                _query_list.append(
                    {
                        "field": "classification_id",
                        "type": "str[]",
                        "value": classification_ids,
                    }
                )
            model_ids = organization_permission_data["inst_names"]
            if model_ids:
                _query_list.append({"field": "model_id", "type": "str[]", "value": model_ids})
                if creator:
                    # 只有创建人条件
                    _query_list.append({"field": "_creator", "type": "str=", "value": creator})

            format_permission_dict[organization_id] = _query_list

        with GraphClient() as ag:
            query = dict(
                label=MODEL,
                params=[],
                order=order,
                order_type=order_type,
                format_permission_dict=format_permission_dict,
                param_type="OR",
                organization_field="group",
            )
            models, _ = ag.query_entity(**query)
            if not include_hidden:
                visible_models = BusinessModelVisibility.filter_models(
                    models,
                    graph=ag,
                )
                models = [model for model in models if model["model_id"] in visible_models]

        lan = SettingLanguage(language)

        for model in models:
            model["model_name"] = lan.get_val("MODEL", model["model_id"]) or model["model_name"]
            # 确保所有模型都有order_id
            if "order_id" not in model:
                model["order_id"] = 0
            if "is_visible" not in model:
                model["is_visible"] = True

        return models

    @staticmethod
    def parse_attrs(attrs: str):
        return json.loads(attrs.replace('\\"', '"'))

    @staticmethod
    def create_model_attr(model_id, attr_info, username="admin"):
        """
        创建模型属性
        """
        with GraphClient() as ag:
            if attr_info.get("attr_type") == "tag":
                attr_info["attr_id"] = TAG_ATTR_ID
                attr_info["editable"] = True
                attr_info["is_only"] = False
                attr_info["is_required"] = False
                attr_info["option"] = ModelManage.normalize_tag_field_option(attr_info.get("option") or {})

            if attr_info.get("attr_type") == "enum":
                ModelManage.normalize_enum_public_binding(attr_info)
                ModelManage.ensure_enum_select_mode(attr_info)

            attr_info = ModelManage.sanitize_attr_default_value(attr_info, log_context="create_model_attr")

            # 企业版字段类型规则（附件/图片：强制可选、非唯一、无约束旋钮）
            enterprise_ext = get_model_enterprise_extension()
            attr_info = enterprise_ext.validate_attr(attr_info)

            ModelManage._validate_attr_id(attr_info["attr_id"])
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, _ = ag.query_entity(MODEL, [model_query])
            model_count = len(models)
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
            ModelManage.validate_tag_attr_definition(attrs, attr_info)
            if attr_info["attr_id"] in {i["attr_id"] for i in attrs}:
                raise BaseAppException("model attr repetition")
            attrs.append(attr_info)

            ModelManage._add_display_field_to_attrs(attrs, attr_info, model_id)

            result = ag.set_entity_properties(MODEL, [model_info["_id"]], dict(attrs=json.dumps(attrs)), {}, [], False)

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        updated_attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))
        ExcludeFieldsCache.update_on_model_change(model_id)

        attrs = updated_attrs

        attr = None
        for attr in attrs:
            if attr["attr_id"] != attr_info["attr_id"]:
                continue
            break

        change_message = f"创建模型属性. 模型名称: {model_info['model_name']}"
        enterprise_message = enterprise_ext.build_attr_change_message({}, attr or attr_info)
        if enterprise_message:
            change_message = f"{change_message}. {enterprise_message}"

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=CREATE_INST,
            message=change_message,
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
            scenario=MODEL_MANAGEMENT_CHANGE,
        )

        return attr

    @staticmethod
    def update_model_attr(model_id, attr_info, username="admin"):
        """
        更新模型属性
        """
        ModelManage._guard_protected_model_attr(attr_info["attr_id"], "修改")
        with GraphClient() as ag:
            ModelManage._validate_attr_id(attr_info["attr_id"])
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, model_count = ag.query_entity(MODEL, [model_query])
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
            if attr_info["attr_id"] not in {i["attr_id"] for i in attrs}:
                raise BaseAppException("model attr not present")

            current_attr = next(
                (attr for attr in attrs if attr["attr_id"] == attr_info["attr_id"]),
                None,
            )
            if not current_attr:
                raise BaseAppException("model attr not present")

            # 附件/图片字段类型创建后不可切换
            if current_attr.get("attr_type") != attr_info.get("attr_type") and (
                is_file_attr_type(current_attr.get("attr_type")) or is_file_attr_type(attr_info.get("attr_type"))
            ):
                raise BaseAppException("附件/图片字段类型创建后不可切换")

            # 企业版字段类型规则（附件/图片：强制可选、非唯一、无约束旋钮）
            enterprise_ext = get_model_enterprise_extension()
            attr_info = enterprise_ext.validate_attr(attr_info)
            current_attr_snapshot = dict(current_attr)

            if current_attr.get("is_required") != attr_info.get("is_required"):
                guard_attr_change_against_unique_rules(
                    model_id,
                    attr_info["attr_id"],
                    attr_info,
                    "update_required",
                    username,
                )
            if current_attr.get("attr_type") != attr_info.get("attr_type"):
                guard_attr_change_against_unique_rules(
                    model_id,
                    attr_info["attr_id"],
                    attr_info,
                    "update_type",
                    username,
                )

            is_tag_attr = ModelManage._is_tag_attr(current_attr)
            if is_tag_attr:
                if attr_info.get("attr_id") != TAG_ATTR_ID:
                    raise BaseAppException("tag 字段 attr_id 不允许修改")
                attr_info["attr_type"] = "tag"
                attr_info["editable"] = True
                attr_info["is_only"] = False
                attr_info["is_required"] = False
                attr_info["option"] = ModelManage.normalize_tag_field_option(attr_info.get("option") or current_attr.get("option") or {})

            is_enum_attr = current_attr.get("attr_type") == "enum"
            if is_enum_attr:
                ModelManage.validate_enum_rule_immutable(current_attr, attr_info)
                ModelManage.validate_enum_select_mode_immutable(current_attr, attr_info)
                ModelManage.normalize_enum_public_binding(attr_info, current_attr)

            attr_info = ModelManage.sanitize_attr_default_value(attr_info, log_context="update_model_attr")

            for attr in attrs:
                if attr_info["attr_id"] != attr["attr_id"]:
                    continue
                attr.update(
                    attr_group=attr_info["attr_group"],
                    attr_name=attr_info["attr_name"],
                    is_required=False if is_tag_attr else attr_info["is_required"],
                    editable=True if is_tag_attr else attr_info["editable"],
                    option=attr_info["option"],
                    user_prompt=attr_info["user_prompt"],
                    default_value=attr_info.get("default_value", []),
                )
                if "governance" in attr_info:
                    attr["governance"] = attr_info["governance"]
                if is_enum_attr:
                    attr["enum_rule_type"] = attr_info.get("enum_rule_type", "custom")
                    attr["public_library_id"] = attr_info.get("public_library_id")
                    attr["enum_select_mode"] = current_attr.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT)

            result = ag.set_entity_properties(MODEL, [model_info["_id"]], dict(attrs=json.dumps(attrs)), {}, [], False)

        attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(model_id)

        attr = None
        for attr in attrs:
            if attr["attr_id"] == attr_info["attr_id"]:
                attr = attr
                break

        change_message = f"修改模型属性. 模型名称: {model_info['model_name']}"
        enterprise_message = enterprise_ext.build_attr_change_message(current_attr_snapshot, attr or attr_info)
        if enterprise_message:
            change_message = f"{change_message}. {enterprise_message}"

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=UPDATE_INST,
            message=change_message,
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
            scenario=MODEL_MANAGEMENT_CHANGE,
        )

        return attr

    @staticmethod
    def update_enum_instances_display(model_id: str, attr_id: str, new_options: list):
        """
        更新枚举类型字段变更后所有实例的 _display 冗余字段
        Args:
            model_id: 模型ID
            attr_id: 属性ID
            new_options: 新的枚举选项列表
        Returns:
            int: 更新的实例数量
        """
        from apps.cmdb.display_field import DisplayFieldConverter

        updated_count = 0
        display_field_id = f"{attr_id}_display"
        checkpoint_key, options_digest = ModelManage._enum_display_backfill_checkpoint(model_id, attr_id, new_options)

        try:
            with GraphClient() as ag:
                scan_cursor, checkpoint_available = ModelManage._get_enum_display_backfill_checkpoint(
                    checkpoint_key,
                    options_digest,
                )
                if not checkpoint_available:
                    raise RuntimeError("枚举显示字段回填断点不可用，已安全跳过本轮图写")
                property_values = []
                while True:
                    query_params = [{"field": "model_id", "type": "str=", "value": model_id}]
                    if scan_cursor is not None:
                        query_params.append(
                            {
                                "field": "id",
                                "type": "id>",
                                "value": scan_cursor,
                            }
                        )
                    instances, _ = ag.query_entity(
                        INSTANCE,
                        query_params,
                        page={"skip": 0, "limit": MAX_BATCH_UPDATE_PROPERTY_VALUES},
                        include_count=False,
                    )
                    if not instances:
                        if property_values:
                            updated_count += ModelManage._write_enum_display_batch(
                                ag,
                                display_field_id,
                                property_values,
                            )
                            checkpoint_available = (
                                ModelManage._save_enum_display_backfill_checkpoint(
                                    checkpoint_key,
                                    options_digest,
                                    property_values[-1]["id"],
                                )
                                and checkpoint_available
                            )
                        ModelManage._clear_enum_display_backfill_checkpoint(checkpoint_key)
                        break

                    next_instance_id = int(instances[-1]["_id"])
                    if scan_cursor is not None and next_instance_id <= scan_cursor:
                        raise RuntimeError(f"枚举显示字段回填游标未推进: cursor={scan_cursor}, next_cursor={next_instance_id}")

                    for instance in instances:
                        if attr_id not in instance or not instance[attr_id]:
                            continue
                        property_values.append(
                            {
                                "id": instance["_id"],
                                "value": DisplayFieldConverter.convert_enum(instance[attr_id], new_options),
                            }
                        )
                        if len(property_values) == MAX_BATCH_UPDATE_PROPERTY_VALUES:
                            updated_count += ModelManage._write_enum_display_batch(
                                ag,
                                display_field_id,
                                property_values,
                            )
                            checkpoint_available = (
                                ModelManage._save_enum_display_backfill_checkpoint(
                                    checkpoint_key,
                                    options_digest,
                                    property_values[-1]["id"],
                                )
                                and checkpoint_available
                            )
                            property_values = []

                    scan_cursor = next_instance_id
                    if not property_values:
                        checkpoint_available = (
                            ModelManage._save_enum_display_backfill_checkpoint(
                                checkpoint_key,
                                options_digest,
                                scan_cursor,
                            )
                            and checkpoint_available
                        )

                if not checkpoint_available:
                    logger.warning(
                        "[update_enum_instances_display] 断点写入不可用，本轮写入保持幂等，后续请求将从头执行: 模型=%s, 字段=%s",
                        model_id,
                        attr_id,
                    )

                if updated_count > 0:
                    logger.info(
                        f"[update_enum_instances_display] 枚举选项变更，已更新 {updated_count} 个实例的 {display_field_id} 字段, "
                        f"模型: {model_id}, 字段: {attr_id}"
                    )

        except Exception as e:
            logger.error(
                f"[update_enum_instances_display] 更新实例枚举 _display 字段失败: 模型={model_id}, 字段={attr_id}, 错误={e}",
                exc_info=True,
            )
            raise BaseAppException("枚举属性已更新，但实例展示字段刷新失败，请重试保存") from e

        return updated_count

    @staticmethod
    def rebuild_file_instances_display(model_id: str, attr_id: str):
        """回填/重算附件/图片字段所有实例的 _display 冗余字段（文件名词干）。

        用于模型字段变更时确保历史实例也有可被全文检索命中的冗余字段。
        文件名词干由实例自身的文件数据决定（不依赖模型配置），故重算即幂等回填。

        Args:
            model_id: 模型ID
            attr_id: 文件字段属性ID

        Returns:
            int: 回填/更新的实例数量
        """
        from apps.cmdb.display_field import DisplayFieldConverter

        updated_count = 0
        display_field_id = f"{attr_id}_display"

        try:
            with GraphClient() as ag:
                last_instance_id = None
                property_values = []
                while True:
                    query_params = [{"field": "model_id", "type": "str=", "value": model_id}]
                    if last_instance_id is not None:
                        query_params.append(
                            {
                                "field": "id",
                                "type": "id>",
                                "value": last_instance_id,
                            }
                        )
                    instances, _ = ag.query_entity(
                        INSTANCE,
                        query_params,
                        page={"skip": 0, "limit": MAX_BATCH_UPDATE_PROPERTY_VALUES},
                        include_count=False,
                    )
                    if not instances:
                        break

                    for instance in instances:
                        if attr_id not in instance or not instance[attr_id]:
                            continue
                        property_values.append(
                            {
                                "id": instance["_id"],
                                "value": DisplayFieldConverter.convert_file(instance[attr_id]),
                            }
                        )
                        if len(property_values) == MAX_BATCH_UPDATE_PROPERTY_VALUES:
                            ag.batch_update_node_property_values(
                                INSTANCE,
                                display_field_id,
                                property_values,
                            )
                            updated_count += len(property_values)
                            property_values = []

                    if len(instances) < MAX_BATCH_UPDATE_PROPERTY_VALUES:
                        break
                    last_instance_id = instances[-1]["_id"]

                if property_values:
                    ag.batch_update_node_property_values(
                        INSTANCE,
                        display_field_id,
                        property_values,
                    )
                    updated_count += len(property_values)

                if updated_count > 0:
                    logger.info(
                        f"[rebuild_file_instances_display] 已回填 {updated_count} 个实例的 {display_field_id} 字段, "
                        f"模型: {model_id}, 字段: {attr_id}"
                    )

        except Exception as e:
            logger.error(
                f"[rebuild_file_instances_display] 回填文件 _display 字段失败: 模型={model_id}, 字段={attr_id}, 错误={e}",
                exc_info=True,
            )
            # 不抛出异常，避免中断主流程

        return updated_count

    @staticmethod
    def delete_model_attr(model_id: str, attr_id: str, username: str = "admin"):
        """
        删除模型属性
        """
        ModelManage._guard_protected_model_attr(attr_id, "删除")
        guard_attr_change_against_unique_rules(
            model_id,
            attr_id,
            None,
            "delete",
            username,
        )
        with GraphClient() as ag:
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, model_count = ag.query_entity(MODEL, [model_query])
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))

            fields_to_remove = [attr_id]

            # 检查是否需要同时删除 _display 字段
            for attr in attrs:
                if attr["attr_id"] == attr_id and attr.get("attr_type") in DISPLAY_FIELD_TYPES:
                    display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"
                    fields_to_remove.append(display_field_id)
                    break

            new_attrs = [attr for attr in attrs if attr["attr_id"] not in fields_to_remove]
            result = ag.set_entity_properties(
                MODEL,
                [model_info["_id"]],
                dict(attrs=json.dumps(new_attrs)),
                {},
                [],
                False,
            )

            # 模型属性删除后，要删除对应模型实例的属性(包括 _display 字段)
            model_params = [{"field": "model_id", "type": "str=", "value": model_id}]
            ag.remove_entitys_properties(INSTANCE, model_params, fields_to_remove)

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        updated_attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))
        ExcludeFieldsCache.update_on_model_change(model_id)

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=DELETE_INST,
            message=f"删除模型属性. 模型名称: {model_info['model_name']}",
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
            scenario=MODEL_MANAGEMENT_CHANGE,
        )

        return updated_attrs

    @staticmethod
    def search_model_info(model_id: str):
        """
        查询模型详情
        """
        query_data = [{"field": "model_id", "type": "str=", "value": model_id}]
        with GraphClient() as ag:
            models, _ = ag.query_entity(MODEL, query_data)
        if len(models) == 0:
            return {}

        model = models[0]

        # if not display_field:
        #     return model
        #
        # # 过滤掉 is_display_field 为 true 的字段
        # if "attrs" in model and model["attrs"]:
        #     attrs = ModelManage.parse_attrs(model["attrs"])
        #     filtered_attrs = [attr for attr in attrs if not attr.get("is_display_field")]
        #     model["attrs"] = json.dumps(filtered_attrs, ensure_ascii=False)

        return model

    @staticmethod
    def get_organization_option(items: list, result: list, name_prefix: str = ""):
        for item in items:
            if name_prefix:
                name = f"{name_prefix}/{item['name']}"
            else:
                name = item["name"]
            result.append(
                dict(
                    id=item["id"],
                    name=name,
                    is_default=False,
                    type="str",
                )
            )
            if item["subGroups"]:
                ModelManage.get_organization_option(item["subGroups"], result, name)

    @staticmethod
    def get_cached_organization_options(
        system_mgmt_client_provider: Callable[[], SystemMgmt],
    ) -> list[dict]:
        option = cache.get(MODEL_ATTR_ORGANIZATION_OPTION_CACHE_KEY)
        if option is not None:
            return ModelManage._clone_options(option)

        groups = UserGroup.get_all_groups(system_mgmt_client_provider()) or []
        option = []
        ModelManage.get_organization_option(groups, option)
        cache.set(
            MODEL_ATTR_ORGANIZATION_OPTION_CACHE_KEY,
            option,
            timeout=MODEL_ATTR_OPTION_CACHE_TTL,
        )
        return ModelManage._clone_options(option)

    @staticmethod
    def get_cached_user_options(
        system_mgmt_client_provider: Callable[[], SystemMgmt],
    ) -> list[dict]:
        option = cache.get(MODEL_ATTR_USER_OPTION_CACHE_KEY)
        if option is not None:
            return ModelManage._clone_options(option)

        users = UserGroup.get_all_users(system_mgmt_client_provider())
        option = [
            dict(
                id=user["id"],
                name=user["username"],
                username=user.get("username"),
                display_name=user.get("display_name"),
                is_default=False,
                type="str",
            )
            for user in users["users"]
        ]
        cache.set(
            MODEL_ATTR_USER_OPTION_CACHE_KEY,
            option,
            timeout=MODEL_ATTR_OPTION_CACHE_TTL,
        )
        return ModelManage._clone_options(option)

    @staticmethod
    def search_model_attr(model_id: str, language: str = "en"):
        """
        查询模型属性
        """
        model_info = ModelManage.search_model_info(model_id)
        attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model_info.get("attrs", "[]")))
        attrs = [ModelManage.sanitize_attr_default_value(attr, log_context="search_model_attr") for attr in attrs]
        unique_rules = build_unique_rule_context(model_id).unique_rules
        # TODO 语言包
        # lan = SettingLanguage(language)
        # model_attr = lan.get_val("ATTR", model_id)
        # for attr in attrs:
        #     if model_attr:
        #         attr["attr_name"] = model_attr.get(attr["attr_id"]) or attr["attr_name"]
        #
        return enrich_attrs_with_unique_display(attrs, unique_rules, model_id)

    @staticmethod
    def search_model_attr_v2(model_id: str):
        """
        查询模型属性
        """
        model_info = ModelManage.search_model_info(model_id)
        attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model_info.get("attrs", "[]")))
        attrs = [ModelManage.sanitize_attr_default_value(attr, log_context="search_model_attr_v2") for attr in attrs]
        unique_rules = build_unique_rule_context(model_id).unique_rules
        attr_types = {attr["attr_type"] for attr in attrs}
        system_mgmt_client = None

        def get_system_mgmt_client() -> SystemMgmt:
            nonlocal system_mgmt_client
            if system_mgmt_client is None:
                system_mgmt_client = SystemMgmt()
            return system_mgmt_client

        if ORGANIZATION in attr_types:
            option = ModelManage.get_cached_organization_options(get_system_mgmt_client)
            for attr in attrs:
                if attr["attr_type"] == ORGANIZATION:
                    attr.update(option=option)

        if USER in attr_types:
            option = ModelManage.get_cached_user_options(get_system_mgmt_client)
            for attr in attrs:
                if attr["attr_type"] == USER:
                    attr.update(option=option)

        return enrich_attrs_with_unique_display(attrs, unique_rules, model_id)

    @staticmethod
    def get_model_unique_rules(model_id: str, editing_rule_id: str | None = None):
        return {
            "rules": list_unique_rules(model_id),
            "candidate_fields": [asdict(candidate) for candidate in list_unique_rule_candidate_fields(model_id, editing_rule_id)],
        }

    @staticmethod
    def create_model_unique_rule(model_id: str, data: dict[str, Any], username: str = "admin"):
        create_unique_rule(
            model_id,
            UniqueRulePayload(field_ids=list(data.get("field_ids") or [])),
            username,
        )
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def update_model_unique_rule(
        model_id: str,
        rule_id: str,
        data: dict[str, Any],
        username: str = "admin",
    ):
        update_unique_rule(
            model_id,
            rule_id,
            UniqueRulePayload(field_ids=list(data.get("field_ids") or [])),
            username,
        )
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def delete_model_unique_rule(model_id: str, rule_id: str, username: str = "admin"):
        delete_unique_rule(model_id, rule_id, username)
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def model_association_create(**data):
        """
        创建模型关联
        """
        with GraphClient() as ag:
            try:
                edge = ag.create_edge(
                    MODEL_ASSOCIATION,
                    data["src_id"],
                    MODEL,
                    data["dst_id"],
                    MODEL,
                    data,
                    "model_asst_id",
                )
            except BaseAppException as e:
                if e.message == "edge already exists":
                    raise BaseAppException("model association repetition")
                else:
                    raise BaseAppException(e.message)
        return edge

    @staticmethod
    def model_association_delete(id: int):
        """
        删除模型关联
        """
        association = None
        with GraphClient() as ag:
            association = ag.query_edge_by_id(id)

        with GraphClient() as ag:
            ag.delete_edge(id)

        model_asst_id = str((association or {}).get("model_asst_id") or "").strip()
        if model_asst_id:
            from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

            schedule_rule_auto_relation_full_sync([model_asst_id])

    @staticmethod
    def model_association_info_search(model_asst_id: str):
        """
        查询模型关联详情
        """
        with GraphClient() as ag:
            query_data = {
                "field": "model_asst_id",
                "type": "str=",
                "value": model_asst_id,
            }
            edges = ag.query_edge(MODEL_ASSOCIATION, [query_data])
        if len(edges) == 0:
            return {}
        return edges[0]

    @staticmethod
    def model_association_search(
        model_id: str,
        *,
        business_only: bool = False,
        language: str = "en",
    ):
        """
        查询模型所有的关联
        """
        query_list = [
            {"field": "src_model_id", "type": "str=", "value": model_id},
            {"field": "dst_model_id", "type": "str=", "value": model_id},
        ]
        with GraphClient() as ag:
            edges = ag.query_edge(MODEL_ASSOCIATION, query_list, param_type="OR")

        if business_only:
            return BusinessModelVisibility.filter_associations(
                edges,
                language=language,
            )
        return edges

    @staticmethod
    def get_model_auto_relation_rules(model_id: str):
        associations = ModelManage.model_association_search(model_id)
        result = []
        for association in associations:
            rule_set = parse_auto_relation_rule_set(association.get(AUTO_RELATION_RULE_FIELD))
            if not rule_set:
                continue
            for rule in rule_set.rules:
                result.append(build_auto_relation_rule_response(association, rule))
        return result

    @staticmethod
    def _get_model_attrs_for_auto_rule(model_id: str):
        model_info = ModelManage.search_model_info(model_id)
        if not model_info:
            raise BaseAppException("模型不存在")
        return ModelManage.parse_attrs(model_info.get("attrs", "[]"))

    @staticmethod
    def save_model_auto_relation_rule(
        model_id: str,
        model_asst_id: str,
        payload: dict[str, Any],
        username: str = "admin",
    ):
        association = ModelManage.model_association_info_search(model_asst_id)
        if not association:
            raise BaseAppException("模型关联不存在")
        if model_id not in {association.get("src_model_id"), association.get("dst_model_id")}:
            raise BaseAppException("模型关联不属于当前模型")

        src_attrs = ModelManage._get_model_attrs_for_auto_rule(association["src_model_id"])
        dst_attrs = ModelManage._get_model_attrs_for_auto_rule(association["dst_model_id"])
        validated_rule = validate_auto_relation_rule_payload(
            association,
            src_attrs,
            dst_attrs,
            payload,
        )
        current_rule_set = parse_auto_relation_rule_set(association.get(AUTO_RELATION_RULE_FIELD)) or AutoRelationRuleSet(version=2, rules=[])
        if any(rule.rule_id == validated_rule.rule_id for rule in current_rule_set.rules):
            raise BaseAppException("自动关联规则标识重复")
        persisted_rule = AutoRelationRule(
            rule_id=validated_rule.rule_id or uuid.uuid4().hex,
            enabled=validated_rule.enabled,
            match_pairs=validated_rule.match_pairs,
            updated_by=username,
            updated_at=timezone.now().isoformat(),
        )
        next_rules = list(current_rule_set.rules)
        next_rules.append(persisted_rule)
        persisted_rule_set = AutoRelationRuleSet(version=2, rules=next_rules)

        with GraphClient() as ag:
            ag.set_edge_properties(
                association["_id"],
                {
                    AUTO_RELATION_RULE_FIELD: dump_auto_relation_rule_set(persisted_rule_set),
                },
            )

        from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

        schedule_rule_auto_relation_full_sync([model_asst_id])

        return build_auto_relation_rule_response(association, persisted_rule)

    @staticmethod
    def update_model_auto_relation_rule(
        model_id: str,
        model_asst_id: str,
        rule_id: str,
        payload: dict[str, Any],
        username: str = "admin",
    ):
        association = ModelManage.model_association_info_search(model_asst_id)
        if not association:
            raise BaseAppException("模型关联不存在")
        if model_id not in {association.get("src_model_id"), association.get("dst_model_id")}:
            raise BaseAppException("模型关联不属于当前模型")

        src_attrs = ModelManage._get_model_attrs_for_auto_rule(association["src_model_id"])
        dst_attrs = ModelManage._get_model_attrs_for_auto_rule(association["dst_model_id"])
        validated_rule = validate_auto_relation_rule_payload(
            association,
            src_attrs,
            dst_attrs,
            {**payload, "rule_id": rule_id},
        )

        current_rule_set = parse_auto_relation_rule_set(association.get(AUTO_RELATION_RULE_FIELD))
        if not current_rule_set:
            raise BaseAppException("自动关联规则不存在")

        replaced = False
        next_rules: list[AutoRelationRule] = []
        for current_rule in current_rule_set.rules:
            if current_rule.rule_id != rule_id:
                next_rules.append(current_rule)
                continue
            replaced = True
            next_rules.append(
                AutoRelationRule(
                    rule_id=rule_id,
                    enabled=validated_rule.enabled,
                    match_pairs=validated_rule.match_pairs,
                    updated_by=username,
                    updated_at=timezone.now().isoformat(),
                )
            )

        if not replaced:
            raise BaseAppException("自动关联规则不存在")

        persisted_rule = next((rule for rule in next_rules if rule.rule_id == rule_id), None)
        persisted_rule_set = AutoRelationRuleSet(version=2, rules=next_rules)

        with GraphClient() as ag:
            ag.set_edge_properties(
                association["_id"],
                {
                    AUTO_RELATION_RULE_FIELD: dump_auto_relation_rule_set(persisted_rule_set),
                },
            )

        from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

        schedule_rule_auto_relation_full_sync([model_asst_id])

        return build_auto_relation_rule_response(association, persisted_rule)

    @staticmethod
    def delete_model_auto_relation_rule(model_id: str, model_asst_id: str, rule_id: str):
        association = ModelManage.model_association_info_search(model_asst_id)
        if not association:
            raise BaseAppException("模型关联不存在")
        if model_id not in {association.get("src_model_id"), association.get("dst_model_id")}:
            raise BaseAppException("模型关联不属于当前模型")

        current_rule_set = parse_auto_relation_rule_set(association.get(AUTO_RELATION_RULE_FIELD))
        if not current_rule_set:
            raise BaseAppException("自动关联规则不存在")

        next_rules = [rule for rule in current_rule_set.rules if rule.rule_id != rule_id]
        if len(next_rules) == len(current_rule_set.rules):
            raise BaseAppException("自动关联规则不存在")

        serialized_rule_set = dump_auto_relation_rule_set(AutoRelationRuleSet(version=2, rules=next_rules)) if next_rules else ""

        with GraphClient() as ag:
            ag.set_edge_properties(
                association["_id"],
                {
                    AUTO_RELATION_RULE_FIELD: serialized_rule_set,
                },
            )

        from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

        schedule_rule_auto_relation_full_sync([model_asst_id])

        return True

    @staticmethod
    def replace_model_auto_relation_rule_set(
        model_id: str,
        model_asst_id: str,
        payload: dict[str, Any],
        username: str = "system",
    ):
        association = ModelManage.model_association_info_search(model_asst_id)
        if not association:
            raise BaseAppException("模型关联不存在")
        if model_id not in {association.get("src_model_id"), association.get("dst_model_id")}:
            raise BaseAppException("模型关联不属于当前模型")

        src_attrs = ModelManage._get_model_attrs_for_auto_rule(association["src_model_id"])
        dst_attrs = ModelManage._get_model_attrs_for_auto_rule(association["dst_model_id"])
        validated_rule_set = validate_auto_relation_rule_set_payload(
            association,
            src_attrs,
            dst_attrs,
            payload,
        )

        persisted_rule_set = AutoRelationRuleSet(
            version=int(validated_rule_set.version or 2),
            rules=[
                AutoRelationRule(
                    rule_id=rule.rule_id,
                    enabled=rule.enabled,
                    match_pairs=rule.match_pairs,
                    updated_by=username,
                    updated_at=timezone.now().isoformat(),
                )
                for rule in validated_rule_set.rules
            ],
        )

        with GraphClient() as ag:
            ag.set_edge_properties(
                association["_id"],
                {
                    AUTO_RELATION_RULE_FIELD: dump_auto_relation_rule_set(persisted_rule_set),
                },
            )

        from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

        schedule_rule_auto_relation_full_sync([model_asst_id])

        return persisted_rule_set

    @staticmethod
    def _build_model_asst_id(src_model_id: str, asst_id: str, dst_model_id: str) -> str:
        src_model_id = str(src_model_id or "").strip()
        asst_id = str(asst_id or "").strip()
        dst_model_id = str(dst_model_id or "").strip()
        if not src_model_id or not asst_id or not dst_model_id:
            raise BaseAppException("src_model_id / asst_id / dst_model_id 不能为空")
        return f"{src_model_id}_{asst_id}_{dst_model_id}"

    @staticmethod
    def _parse_auto_relation_rule_set_cell(value, context: str) -> dict[str, Any]:
        try:
            data = value if isinstance(value, dict) else json.loads(str(value))
        except Exception:
            raise BaseAppException(f"{context} auto_relation_rule 不是合法 JSON")

        if not isinstance(data, (dict, list)):
            raise BaseAppException(f"{context} auto_relation_rule 必须是对象或数组")
        data = canonicalize_auto_relation_rule_set_payload(data)
        if not isinstance(data.get("rules"), list) or not data.get("rules"):
            raise BaseAppException(f"{context} auto_relation_rule.rules 不能为空")
        return data

    @staticmethod
    def _validate_auto_rule_sheet_authority(sheet_name: str, src_model_id: str, context: str):
        expected_sheet = f"asso-{src_model_id}"
        if sheet_name != expected_sheet:
            raise BaseAppException(f"{context} 自动关联规则只能在 sheet[{expected_sheet}] 中定义")

    @staticmethod
    def _is_empty_auto_rule_sheet_row(row: dict[str, Any]) -> bool:
        return all(str(row.get(key) or "").strip() == "" for key in ("src_model_id", "dst_model_id", "asst_id", AUTO_RELATION_RULE_FIELD))

    @staticmethod
    def _import_auto_relation_rule_sets_from_asso_sheets(model_config: dict[str, list[dict]]):
        pending_items = []
        seen_model_asst_ids: dict[str, str] = {}

        for sheet_name, rows in model_config.items():
            if not sheet_name.startswith("asso-"):
                continue

            for row_index, row in enumerate(rows, start=3):
                if ModelManage._is_empty_auto_rule_sheet_row(row):
                    continue
                if AUTO_RELATION_RULE_FIELD not in row:
                    continue

                src_model_id = str(row.get("src_model_id", "")).strip()
                dst_model_id = str(row.get("dst_model_id", "")).strip()
                asst_id = str(row.get("asst_id", "")).strip()
                context = f"sheet[{sheet_name}] 第 {row_index} 行"
                raw_rule_set = row.get(AUTO_RELATION_RULE_FIELD, "")

                model_asst_id = ModelManage._build_model_asst_id(src_model_id, asst_id, dst_model_id)

                if raw_rule_set in (None, ""):
                    continue

                ModelManage._validate_auto_rule_sheet_authority(sheet_name, src_model_id, context)

                previous_context = seen_model_asst_ids.get(model_asst_id)
                if previous_context:
                    raise BaseAppException(f"定义了重复的关联规则：{context} 【{model_asst_id}】 与 【{previous_context}】")
                seen_model_asst_ids[model_asst_id] = context

                payload = ModelManage._parse_auto_relation_rule_set_cell(raw_rule_set, context)

                pending_items.append(
                    {
                        "context": context,
                        "model_id": src_model_id,
                        "model_asst_id": model_asst_id,
                        "payload": payload,
                    }
                )

        for item in pending_items:
            try:
                if item["payload"] is None:
                    association = ModelManage.model_association_info_search(item["model_asst_id"])
                    if association and association.get(AUTO_RELATION_RULE_FIELD):
                        with GraphClient() as ag:
                            ag.set_edge_properties(
                                association["_id"],
                                {
                                    AUTO_RELATION_RULE_FIELD: "",
                                },
                            )
                        from apps.cmdb.services.auto_relation_reconcile import schedule_rule_auto_relation_full_sync

                        schedule_rule_auto_relation_full_sync([item["model_asst_id"]])
                else:
                    ModelManage.replace_model_auto_relation_rule_set(
                        item["model_id"],
                        item["model_asst_id"],
                        item["payload"],
                        username="system",
                    )
            except Exception as err:
                raise BaseAppException(f"{item['context']} ({item['model_asst_id']}) 导入自动关联规则失败: {getattr(err, 'message', str(err))}")

    @staticmethod
    def check_model_exist_association(model_id):
        """模型存在关联关系"""
        edges = ModelManage.model_association_search(model_id)
        if edges:
            raise BaseAppException("model association exist")

    @staticmethod
    def check_model_exist_inst(model_id):
        """模型存在实例"""
        params = [{"field": "model_id", "type": "str=", "value": model_id}]
        with GraphClient() as ag:
            _, count = ag.query_entity(INSTANCE, params, page=dict(skip=0, limit=1))
        if count > 0:
            raise BaseAppException("model exist instance")

    # ===========

    @staticmethod
    def get_max_order_id(classification_id: str):
        """
        获取当前最大的 order_id
        Args:
            classification_id: 分类ID
        """
        with GraphClient() as ag:
            models, _ = ag.query_entity(
                MODEL,
                [
                    {
                        "field": "classification_id",
                        "type": "str=",
                        "value": classification_id,
                    }
                ],
                order="order_id",
                order_type="desc",
                page={"skip": 0, "limit": 1},
            )
            if not models:
                return 0
            return models[0].get("order_id", 0)

    @staticmethod
    def update_model_orders(model_orders: list):
        """
        批量更新模型排序及可见性
        Args:
            model_orders: [
                {"model_id": "model_1", "order_id": 1},
                {"model_id": "model_2", "order_id": 2, "is_visible": False},
            ]
            order_id 必填；is_visible 可选，缺省时不修改原有可见性
        """
        with GraphClient() as ag:
            for order_info in model_orders:
                model_query = {
                    "field": "model_id",
                    "type": "str=",
                    "value": order_info["model_id"],
                }
                models, model_count = ag.query_entity(MODEL, [model_query])
                if model_count == 0:
                    continue
                model_info = models[0]
                props: dict = {"order_id": order_info["order_id"]}
                if "is_visible" in order_info:
                    props["is_visible"] = bool(order_info["is_visible"])
                ag.set_entity_properties(
                    MODEL,
                    [model_info["_id"]],
                    props,
                    {},
                    [],
                    False,
                )
        return True

    @staticmethod
    def export_model_config(language, model_ids=None):
        """导出模型配置为Excel (按model_config.xlsx模板格式)

        :param model_ids: 勾选导出的模型ID列表。为空(None 或空列表)时不过滤,
            导出全量模型配置(向后兼容)。有值时:models/attr 只保留选中模型;
            classifications 只保留选中模型所属的分类;public_enum_libraries 全量不过滤;
            asso 逐行过滤,只保留 src 和 dst 都在选中集合内的关联行,过滤后无行则不生成该 asso sheet。
        """
        from io import BytesIO

        selected_ids = set(model_ids) if model_ids else None

        from openpyxl import Workbook

        CLASSIFICATION_HEADERS_CN = ["模型分类ID", "模型分类名称"]
        CLASSIFICATION_HEADERS_EN = ["classification_id", "classification_name"]

        MODEL_HEADERS_CN = ["模型ID", "模型名称", "模型图标", "模型分类ID"]
        MODEL_HEADERS_EN = ["model_id", "model_name", "icn", "classification_id"]

        ATTR_HEADERS_CN = [
            "英文名",
            "名称",
            "类型",
            "数据配置",
            "分组",
            "是否唯一",
            "是否可编辑",
            "是否必填",
            "用户提示",
            "唯一规则序号",
            "默认值",
        ]
        ATTR_HEADERS_EN = [
            "attr_id",
            "attr_name",
            "attr_type",
            "option",
            "attr_group",
            "is_only",
            "editable",
            "is_required",
            "user_prompt",
            "unique_rule_order",
            "default_value",
        ]
        enterprise_ext = get_model_enterprise_extension()
        extra_attr_headers_cn, extra_attr_headers_en = enterprise_ext.extra_export_attr_headers()
        ATTR_HEADERS_CN.extend(extra_attr_headers_cn)
        ATTR_HEADERS_EN.extend(extra_attr_headers_en)

        ASSO_HEADERS_CN = ["源模型", "目标模型", "关联关系", "源-目标约束", "自动关联规则"]
        ASSO_HEADERS_EN = ["src_model_id", "dst_model_id", "asst_id", "mapping", "auto_relation_rule"]
        PUBLIC_ENUM_LIBRARY_HEADERS_CN = [
            "公共选项库ID",
            "公共选项库名称",
            "组织列表",
            "选项列表",
        ]
        PUBLIC_ENUM_LIBRARY_HEADERS_EN = [
            "library_id",
            "name",
            "team",
            "options",
        ]

        workbook = Workbook()

        ws_classifications = workbook.active
        if ws_classifications is None:
            ws_classifications = workbook.create_sheet(title="classifications")
        else:
            ws_classifications.title = "classifications"
        ws_classifications.append(CLASSIFICATION_HEADERS_CN)
        ws_classifications.append(CLASSIFICATION_HEADERS_EN)

        classifications = ClassificationManage.search_model_classification(language=language)

        models = ModelManage.search_model(language=language)
        if selected_ids is not None:
            models = [m for m in models if m.get("model_id", "") in selected_ids]
        selected_classification_ids = {m.get("classification_id", "") for m in models}

        for classification in classifications:
            cls_id = classification.get("classification_id", "")
            if selected_ids is not None and cls_id not in selected_classification_ids:
                continue
            ws_classifications.append(
                [
                    classification.get("classification_id", ""),
                    classification.get("classification_name", ""),
                ]
            )

        ws_models = workbook.create_sheet(title="models")
        ws_models.append(MODEL_HEADERS_CN)
        ws_models.append(MODEL_HEADERS_EN)

        ws_public_enum_libraries = workbook.create_sheet(title="public_enum_libraries")
        ws_public_enum_libraries.append(PUBLIC_ENUM_LIBRARY_HEADERS_CN)
        ws_public_enum_libraries.append(PUBLIC_ENUM_LIBRARY_HEADERS_EN)

        from apps.cmdb.services.public_enum_library import list_libraries

        for library in list_libraries():
            ws_public_enum_libraries.append(
                [
                    library.get("library_id", ""),
                    library.get("name", ""),
                    json.dumps(library.get("team", []), ensure_ascii=False),
                    json.dumps(library.get("options", []), ensure_ascii=False),
                ]
            )

        # 快速模型（自定义上报）不参与模型管理的导入导出，避免污染正式模型的跨环境流转
        models = [m for m in models if not m.get("is_custom_reporting")]
        for model in models:
            ws_models.append(
                [
                    model.get("model_id", ""),
                    model.get("model_name", ""),
                    model.get("icn", ""),
                    model.get("classification_id", ""),
                ]
            )

        for model in models:
            model_id = model.get("model_id", "")
            if not model_id:
                continue

            ws_attr = workbook.create_sheet(title=f"attr-{model_id}")
            ws_attr.append(ATTR_HEADERS_CN)
            ws_attr.append(ATTR_HEADERS_EN)

            attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model.get("attrs", "[]")))
            unique_rules = build_unique_rule_context(model_id).unique_rules
            attr_rows = []
            for attr in attrs:
                if attr.get("is_display_field"):
                    continue
                option = attr.get("option", {})
                if attr.get("attr_type") == "enum":
                    if attr.get("enum_rule_type") == "public_library":
                        option = {
                            "enum_rule_type": "public_library",
                            "public_library_id": attr.get("public_library_id"),
                            "enum_select_mode": attr.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT),
                        }
                    elif isinstance(option, list):
                        option = option
                option_str = json.dumps(option, ensure_ascii=False) if option else ""
                attr_rows.append(
                    {
                        "attr_id": attr.get("attr_id", ""),
                        "attr_name": attr.get("attr_name", ""),
                        "attr_type": attr.get("attr_type", ""),
                        "option": option_str,
                        "attr_group": attr.get("attr_group", ""),
                        "is_only": attr.get("is_only", False),
                        "editable": attr.get("editable", True),
                        "is_required": attr.get("is_required", False),
                        "user_prompt": attr.get("user_prompt", ""),
                        "default_value": json.dumps(attr.get("default_value", []), ensure_ascii=False) if attr.get("attr_type") == "enum" else "",
                        **enterprise_ext.extend_export_attr_row(attr),
                    }
                )

            attr_rows = apply_unique_rules_to_attr_export_rows(attr_rows, unique_rules)
            logger.info(
                "[UniqueRule] attr export complete model_id=%s sheet_name=%s rule_count=%s",
                model_id,
                f"attr-{model_id}",
                len(unique_rules),
            )
            for row in attr_rows:
                ws_attr.append([row.get(header, "") for header in ATTR_HEADERS_EN])

            associations = ModelManage.model_association_search(
                model_id,
                business_only=True,
                language=language,
            )
            if associations:
                if selected_ids is not None:
                    associations = [
                        a for a in associations if a.get("src_model_id", "") in selected_ids and a.get("dst_model_id", "") in selected_ids
                    ]
                if associations:
                    ws_asso = workbook.create_sheet(title=f"asso-{model_id}")
                    ws_asso.append(ASSO_HEADERS_CN)
                    ws_asso.append(ASSO_HEADERS_EN)
                    for asso in associations:
                        rule_set_value = ""
                        if asso.get("src_model_id") == model_id:
                            rule_set = parse_auto_relation_rule_set(asso.get(AUTO_RELATION_RULE_FIELD))
                            rule_set_value = dump_auto_relation_rule_set_compact(rule_set) if rule_set else ""
                        ws_asso.append(
                            [
                                asso.get("src_model_id", ""),
                                asso.get("dst_model_id", ""),
                                asso.get("asst_id", ""),
                                asso.get("mapping", ""),
                                rule_set_value,
                            ]
                        )

        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        return file_stream

    @staticmethod
    def _apply_model_config_post_import_extras(
        model_config: dict[str, list[dict]],
        *,
        keep_existing_unique_rules_on_conflict: bool = False,
    ):
        with GraphClient() as ag:
            for sheet_name, rows in model_config.items():
                if not sheet_name.startswith("attr-"):
                    continue
                model_id = sheet_name.replace("attr-", "", 1)
                model_info = ModelManage.search_model_info(model_id)
                if not model_info:
                    continue
                try:
                    rules = build_unique_rules_from_attr_rows(model_id, rows)
                    exist_items, _ = ag.query_entity(
                        INSTANCE,
                        [{"field": "model_id", "type": "str=", "value": model_id}],
                    )
                    conflicts = validate_unique_rules_against_existing_instances(
                        model_id,
                        rules,
                        exist_items,
                    )
                    if conflicts:
                        if keep_existing_unique_rules_on_conflict:
                            logger.warning(
                                "[UniqueRule] 存量实例与待应用规则冲突，保留原唯一规则并继续初始化 "
                                "model_id=%s sheet_name=%s conflict_count=%s reason=%s",
                                model_id,
                                sheet_name,
                                len(conflicts),
                                conflicts[0].message,
                            )
                            continue
                        raise BaseAppException(conflicts[0].message)
                    ag.set_entity_properties(
                        MODEL,
                        [model_info["_id"]],
                        {"unique_rules": json.dumps([asdict(rule) for rule in rules], ensure_ascii=False)},
                        {},
                        [],
                        False,
                    )
                    logger.info(
                        "[UniqueRule] attr import success model_id=%s sheet_name=%s rule_count=%s",
                        model_id,
                        sheet_name,
                        len(rules),
                    )
                except Exception as err:
                    logger.warning(
                        "[UniqueRule] attr import failed model_id=%s sheet_name=%s row_number=%s reason=%s",
                        model_id,
                        sheet_name,
                        0,
                        getattr(err, "message", str(err)),
                    )
                    raise

        ModelManage._import_auto_relation_rule_sets_from_asso_sheets(model_config)
        # push/ingest 依赖可写 node_id / monitor_id；模型导入后对一期支持模型幂等补齐
        from apps.cmdb.services.module_ingest import SUPPORTED_INGEST_MODELS, ensure_model_monitor_id_attr, ensure_model_node_id_attr

        for mid in sorted(SUPPORTED_INGEST_MODELS):
            ensure_model_node_id_attr(mid, username="admin")
            ensure_model_monitor_id_attr(mid, username="admin")

    @staticmethod
    def import_model_config(file):
        from apps.cmdb.model_migrate.migrete_service import ModelMigrate

        migrator = ModelMigrate(file_source=file, is_pre=False)
        result = migrator.main()
        ModelManage._apply_model_config_post_import_extras(migrator.model_config)
        return result
