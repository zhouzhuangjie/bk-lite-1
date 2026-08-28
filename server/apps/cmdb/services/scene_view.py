"""跨模型标签场景的活查询、可见范围与导出。"""

from __future__ import annotations

import os
import re
from collections.abc import Callable, Mapping, Sequence
from io import BytesIO

from django.db.models import Q

from apps.cmdb.constants.field_constraints import TAG_ATTR_ID
from apps.cmdb.models.scene_view import SceneView
from apps.cmdb.services.instance import InstanceManage

TagMatch = str
PermissionMapLoader = Callable[[str], dict]
InstanceListFn = Callable[..., tuple[list, int]]

_ORG_SHARE_PERMISSION = "asset_views_scene-Org Share"
_PAGE_SIZE = 200
_MAX_INST_NAME_SEARCH = 128
_MAX_LIST_LEN = 50
_FIELD_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]{0,63}$")
_ALLOWED_SEARCH_TYPES = frozenset(
    {
        "str=",
        "str*",
        "int=",
        "bool",
        "time",
        "list[]",
        "list_any[]",
    }
)


def build_tag_query_list(tags: Sequence[str], tag_match: TagMatch) -> list[dict]:
    cleaned = [str(item).strip() for item in tags if str(item).strip()]
    if not cleaned:
        return []
    if tag_match == "or":
        return [
            {
                "field": TAG_ATTR_ID,
                "type": "list_any[]",
                "value": cleaned,
                "accurate": True,
            }
        ]
    return [
        {
            "field": TAG_ATTR_ID,
            "type": "list_any[]",
            "value": [item],
            "accurate": True,
        }
        for item in cleaned
    ]


def build_inst_name_search(keyword: object) -> dict | None:
    if not isinstance(keyword, str):
        return None
    text = keyword.strip()
    if not text:
        return None
    return {"field": "inst_name", "type": "str*", "value": text[:_MAX_INST_NAME_SEARCH]}


def build_model_search_clause(raw: object) -> dict | None:
    if isinstance(raw, str):
        return build_inst_name_search(raw)
    if not isinstance(raw, dict):
        return None
    field = str(raw.get("field") or "").strip()
    if not _FIELD_RE.match(field):
        return None
    clause_type = str(raw.get("type") or "").strip()
    if clause_type not in _ALLOWED_SEARCH_TYPES:
        return None
    if clause_type == "time":
        start = _bounded_text(raw.get("start"))
        end = _bounded_text(raw.get("end"))
        if not start or not end:
            return None
        return {"field": field, "type": "time", "start": start, "end": end}
    if clause_type in {"list[]", "list_any[]"}:
        values = _bounded_list(raw.get("value"))
        if not values:
            return None
        clause = {"field": field, "type": clause_type, "value": values}
        if raw.get("accurate") is True:
            clause["accurate"] = True
        return clause
    if clause_type == "bool":
        if raw.get("value") not in (True, False):
            return None
        return {"field": field, "type": "bool", "value": raw["value"]}
    if clause_type == "int=":
        number = _as_int(raw.get("value"))
        if number is None:
            return None
        return {"field": field, "type": "int=", "value": number}
    text = _bounded_text(raw.get("value"))
    if not text:
        return None
    return {"field": field, "type": clause_type, "value": text}


def _bounded_text(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    return text[:_MAX_INST_NAME_SEARCH]


def _bounded_list(value: object) -> list:
    if not isinstance(value, list):
        return []
    cleaned = []
    for item in value[:_MAX_LIST_LEN]:
        if isinstance(item, bool) or item is None:
            continue
        if isinstance(item, int):
            cleaned.append(item)
            continue
        if isinstance(item, str):
            text = item.strip()
            if text:
                cleaned.append(text[:_MAX_INST_NAME_SEARCH])
    return cleaned


def _as_int(value: object) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().lstrip("-").isdigit():
        return int(value.strip())
    return None


def user_org_ids(user) -> list[int]:
    ids: list[int] = []
    for item in getattr(user, "group_list", None) or []:
        raw = item.get("id") if isinstance(item, dict) else item
        if raw in (None, ""):
            continue
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    return ids


def _permission_set(user) -> set:
    perms = getattr(user, "permission", None)
    if isinstance(perms, dict):
        return set(perms.get("cmdb", set()) or [])
    return set(perms or [])


def can_publish_org(user) -> bool:
    if getattr(user, "is_superuser", False):
        return True
    return _ORG_SHARE_PERMISSION in _permission_set(user)


def can_publish_global(user) -> bool:
    if getattr(user, "is_superuser", False):
        return True
    roles = set(getattr(user, "roles", None) or [])
    names = {"admin", "cmdb_admin"}
    client_id = os.getenv("CLIENT_ID", "")
    if client_id:
        names.add(f"{client_id}_admin")
    return bool(roles & names)


def can_edit_scene(user, scene: SceneView) -> bool:
    if scene.created_by == getattr(user, "username", ""):
        return True
    return scene.visibility == SceneView.Visibility.GLOBAL and can_publish_global(user)


def build_visible_scene_query(*, username: str, domain: str, org_ids: Sequence[int]) -> Q:
    personal = Q(
        visibility=SceneView.Visibility.PERSONAL,
        created_by=username,
        domain=domain or "",
    )
    org = Q(visibility=SceneView.Visibility.ORGANIZATION, organization__in=list(org_ids)) if org_ids else Q(pk__in=[])
    glob = Q(visibility=SceneView.Visibility.GLOBAL)
    return personal | org | glob


def execute_scene_query(
    *,
    model_ids: Sequence[str],
    tags: Sequence[str],
    tag_match: TagMatch,
    creator: str,
    page: int = 1,
    page_size: int = 20,
    pagination: Mapping[str, tuple[int, int]] | None = None,
    searches: Mapping[str, object] | None = None,
    permission_map_loader: PermissionMapLoader,
    instance_list_fn: InstanceListFn | None = None,
) -> dict:
    list_fn = instance_list_fn or InstanceManage.instance_list
    tag_params = build_tag_query_list(tags, tag_match)
    pager = pagination or {}
    search_map = searches or {}
    models: list[dict] = []
    total = 0
    for model_id in model_ids:
        model_page, model_page_size = pager.get(model_id, (page, page_size))
        search_clause = build_model_search_clause(search_map.get(model_id))
        params = list(tag_params)
        if search_clause:
            params.append(search_clause)
        insts, count = list_fn(
            model_id=model_id,
            params=params,
            page=model_page,
            page_size=model_page_size,
            order="",
            permission_map=permission_map_loader(model_id),
            creator=creator,
            case_sensitive=search_clause.get("type") != "str*" if search_clause else True,
        )
        if not count and not search_clause:
            continue
        models.append({"model_id": model_id, "count": count, "insts": insts})
        total += count
    return {"total": total, "models": models}


def collect_all_scene_instances(
    *,
    model_ids: Sequence[str],
    tags: Sequence[str],
    tag_match: TagMatch,
    creator: str,
    permission_map_loader: PermissionMapLoader,
    instance_list_fn: InstanceListFn | None = None,
    page_size: int = _PAGE_SIZE,
) -> dict:
    list_fn = instance_list_fn or InstanceManage.instance_list
    tag_params = build_tag_query_list(tags, tag_match)
    models: list[dict] = []
    total = 0
    for model_id in model_ids:
        page = 1
        collected: list = []
        count = 0
        while True:
            insts, count = list_fn(
                model_id=model_id,
                params=list(tag_params),
                page=page,
                page_size=page_size,
                order="",
                permission_map=permission_map_loader(model_id),
                creator=creator,
            )
            collected.extend(insts or [])
            if not insts or len(collected) >= count:
                break
            page += 1
        if not count:
            continue
        models.append({"model_id": model_id, "count": count, "insts": collected})
        total += count
    return {"total": total, "models": models}


def merge_model_workbooks(sheets: Sequence[tuple[str, BytesIO]]) -> BytesIO:
    from openpyxl import Workbook, load_workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for raw_name, payload in sheets:
        name = _sheet_name(raw_name)
        source = load_workbook(payload)
        src_sheet = source.active
        dst = workbook.create_sheet(title=name)
        for row in src_sheet.iter_rows(values_only=True):
            dst.append(list(row))
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _sheet_name(value: str) -> str:
    cleaned = "".join(ch for ch in str(value) if ch not in r"[]:*?/\\")
    return (cleaned or "sheet")[:31]
