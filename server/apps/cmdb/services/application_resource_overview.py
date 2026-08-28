import ast
import json
from collections import deque
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook

from apps.cmdb.display_field.constants import DISPLAY_FIELD_TYPES, DISPLAY_SUFFIX
from apps.cmdb.display_field.handler import DisplayFieldHandler
from apps.cmdb.model_ops.extensions import is_file_attr_type
from apps.cmdb.services.instance import InstanceManage
from apps.cmdb.services.model import ModelManage
from apps.cmdb.validators.field_validator import normalize_enum_values
from apps.rpc.node_mgmt import NodeMgmt

APPLICATION_MODELS = {"application", "system"}
DATABASE_MODELS = {
    "mysql",
    "oracle",
    "mssql",
    "redis",
    "mongodb",
    "es",
    "postgresql",
    "db2",
    "tidb",
    "dameng",
    "hbase",
    "opengauss",
    "kingbase",
    "vastbase",
    "greenplum",
    "gbase8a",
    "influxdb",
    "aliyun_mysql",
    "aliyun_pgsql",
    "aliyun_redis",
    "aliyun_mongodb",
    "qcloud_mysql",
    "qcloud_redis",
    "qcloud_mongodb",
    "qcloud_pgsql",
    "aws_rds",
    "aws_docdb",
    "aws_memdb",
    "aws_elasticache",
    "azure_mysql",
    "azure_redis",
    "hwcloud_rds",
    "hwcloud_dcs",
}
MIDDLEWARE_MODELS = {
    "apache",
    "tomcat",
    "nginx",
    "iis",
    "weblogic",
    "websphere",
    "tongweb",
    "keepalive",
    "tuxedo",
    "jetty",
    "openresty",
    "jboss",
    "haproxy",
    "minio",
}
CACHE_MODELS = {"redis", "memcached", "aliyun_redis", "qcloud_redis", "aws_elasticache", "azure_redis", "hwcloud_dcs"}
MESSAGE_QUEUE_MODELS = {
    "kafka",
    "zookeeper",
    "rabbitmq",
    "activemq",
    "etcd",
    "rocketmq",
    "spark",
    "aliyun_kafka_inst",
    "qcloud_rocketmq",
    "qcloud_cmq",
    "qcloud_cmq_topic",
    "qcloud_plusar_cluster",
    "aws_msk",
}
HOST_MODELS = {
    "host",
    "vmware_vm",
    "aliyun_ecs",
    "qcloud_cvm",
    "hwcloud_ecs",
    "aws_ec2",
    "openstack_vm",
    "smartx_vm",
    "fusioncompute_vm",
    "manageone_server",
    "sangforhci_vm",
    "inspurincloudrail_vm",
    "sangforscp_vm",
    "nutanixhci_vm",
}
HARDWARE_MODELS = {
    "physcial_server",
    "disk",
    "memory",
    "nic",
    "gpu",
    "storage",
    "storage_pool",
    "storage_disk",
    "storage_volume",
    "switch",
    "router",
    "firewall",
    "loadbalance",
}
RACK_ROOM_MODELS = {"rack", "server_room", "datacenter"}


class ApplicationResourceOverviewService:
    RESOURCE_GROUP_ORDER = [
        "application",
        "host",
        "database",
        "middleware",
        "cache",
        "message_queue",
        "hardware",
        "rack_room",
        "other",
    ]

    INSTANCE_IDENTITY_FIELDS = ["inst_uuid", "model_id", "inst_name"]
    INSTANCE_EXPORT_CORE_TITLES = {
        "inst_name": "实例名称",
        "model_id": "模型ID",
    }
    _cloud_region_name_cache: dict[str, str] | None = None

    @staticmethod
    def _external_id(instance: dict) -> str | None:
        inst_uuid = instance.get("inst_uuid")
        if inst_uuid in (None, ""):
            return None
        return str(inst_uuid)

    @staticmethod
    def _normalize_node(instance: dict, hop: int = 0) -> dict | None:
        external_id = ApplicationResourceOverviewService._external_id(instance)
        if not external_id:
            return None
        return {
            "id": external_id,
            "name": instance.get("inst_name") or external_id,
            "model_id": instance.get("model_id", ""),
            "hop": hop,
            "category": ApplicationResourceOverviewService._category_for_model(instance.get("model_id", "")),
        }

    @staticmethod
    def _filter_visible_instances(instances: list[dict], permission_map: dict | None, user=None) -> list[dict]:
        if not instances:
            return []
        if not permission_map:
            return [instance for instance in instances if instance.get("_id") is not None]
        ids = set()
        for instance in instances:
            try:
                ids.add(int(instance.get("_id")))
            except (TypeError, ValueError):
                continue
        instances_map = InstanceManage._query_instance_map_by_ids(ids)
        result = []
        for instance in instances:
            try:
                inst_id = int(instance.get("_id"))
            except (TypeError, ValueError):
                continue
            full_instance = instances_map.get(inst_id) or instance
            if InstanceManage._has_topology_view_permission(full_instance, permission_map, user=user):
                result.append(full_instance)
        return result

    @staticmethod
    def list_system_applications(inst_id: int, permission_map: dict | None = None, user=None) -> list[dict]:
        associations = InstanceManage.instance_association_instance_list("system", int(inst_id)) or []
        applications = []
        seen = set()
        for association in associations:
            if association.get("asst_id") != "contains":
                continue
            candidate_model_ids = {association.get("src_model_id"), association.get("dst_model_id")}
            if "application" not in candidate_model_ids:
                continue
            visible_instances = ApplicationResourceOverviewService._filter_visible_instances(
                association.get("inst_list") or [],
                permission_map,
                user=user,
            )
            for instance in visible_instances:
                if instance.get("model_id") != "application":
                    continue
                external_id = ApplicationResourceOverviewService._external_id(instance)
                if not external_id or external_id in seen:
                    continue
                seen.add(external_id)
                applications.append(
                    {
                        "id": external_id,
                        "name": instance.get("inst_name") or external_id,
                        "model_id": "application",
                    }
                )
        return applications

    @staticmethod
    def build_application_topology(
        inst_id: int,
        model_id: str,
        depth: int = 1,
        permission_map: dict | None = None,
        user=None,
    ) -> dict:
        center = InstanceManage.query_entity_by_id(int(inst_id)) or {}
        center_node = ApplicationResourceOverviewService._normalize_node(center, hop=0)
        if not center_node:
            return {"center": {}, "nodes": [], "links": [], "truncated": False}

        nodes = {center_node["id"]: center_node}
        graph_id_to_uuid = {int(inst_id): center_node["id"]}
        links = {}
        visited = {(model_id, int(inst_id))}
        queue = deque([(model_id, int(inst_id), 0)])

        while queue:
            current_model_id, current_inst_id, hop = queue.popleft()
            if hop >= int(depth):
                continue
            current_uuid = graph_id_to_uuid.get(current_inst_id)
            if not current_uuid:
                continue
            associations = InstanceManage.instance_association_instance_list(current_model_id, current_inst_id) or []
            for association in associations:
                other_instances = ApplicationResourceOverviewService._filter_visible_instances(
                    association.get("inst_list") or [],
                    permission_map,
                    user=user,
                )
                for instance in other_instances:
                    try:
                        other_inst_id = int(instance.get("_id"))
                    except (TypeError, ValueError):
                        continue
                    other_node = ApplicationResourceOverviewService._normalize_node(instance, hop=hop + 1)
                    if not other_node:
                        continue
                    other_model_id = instance.get("model_id", "")
                    node_key = other_node["id"]
                    graph_id_to_uuid[other_inst_id] = node_key
                    if node_key not in nodes:
                        nodes[node_key] = other_node
                    else:
                        nodes[node_key]["hop"] = min(nodes[node_key]["hop"], hop + 1)

                    link_key = f"{association.get('model_asst_id')}:{instance.get('inst_asst_id', node_key)}"
                    if link_key not in links:
                        links[link_key] = {
                            "id": link_key,
                            "source": current_uuid,
                            "target": node_key,
                            "asst_id": association.get("asst_id"),
                            "model_asst_id": association.get("model_asst_id"),
                        }

                    visit_key = (other_model_id, other_inst_id)
                    if visit_key not in visited:
                        visited.add(visit_key)
                        queue.append((other_model_id, other_inst_id, hop + 1))

        return {
            "center": center_node,
            "nodes": list(nodes.values()),
            "links": list(links.values()),
            "truncated": False,
        }

    @staticmethod
    def _category_for_model(model_id: str) -> str:
        if model_id in APPLICATION_MODELS:
            return "application"
        if model_id in CACHE_MODELS:
            return "cache"
        if model_id in MESSAGE_QUEUE_MODELS:
            return "message_queue"
        if model_id in DATABASE_MODELS:
            return "database"
        if model_id in MIDDLEWARE_MODELS:
            return "middleware"
        if model_id in HOST_MODELS:
            return "host"
        if model_id in RACK_ROOM_MODELS:
            return "rack_room"
        if model_id in HARDWARE_MODELS:
            return "hardware"
        return "other"

    @staticmethod
    def build_application_resources(inst_id: int, model_id: str, permission_map: dict | None = None, user=None) -> dict:
        topology = ApplicationResourceOverviewService.build_application_topology(
            inst_id=inst_id,
            model_id=model_id,
            depth=3,
            permission_map=permission_map,
            user=user,
        )
        groups: dict[str, list[dict]] = {key: [] for key in ApplicationResourceOverviewService.RESOURCE_GROUP_ORDER}
        for node in topology.get("nodes", []):
            group_key = node.get("category") or ApplicationResourceOverviewService._category_for_model(node.get("model_id", ""))
            groups.setdefault(group_key, []).append(
                {
                    "id": node["id"],
                    "name": node["name"],
                    "model_id": node["model_id"],
                    "hop": node.get("hop", 0),
                }
            )
        groups = {key: value for key, value in groups.items() if value}
        counts = {key: len(value) for key, value in groups.items()}
        return {"groups": groups, "counts": counts}

    @staticmethod
    def _serialize_instance_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item) for item in value if item not in (None, ""))
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        if isinstance(value, str):
            stripped = value.strip()
            if stripped.startswith("[") and stripped.endswith("]"):
                try:
                    parsed = ast.literal_eval(stripped)
                except (ValueError, SyntaxError):
                    parsed = None
                if isinstance(parsed, (list, tuple, set)):
                    return ApplicationResourceOverviewService._serialize_instance_value(parsed)
            if len(stripped) >= 19 and "T" in stripped[:20]:
                try:
                    parsed_dt = datetime.fromisoformat(stripped.replace("Z", "+00:00"))
                    return parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
                except ValueError:
                    return value
            return value
        return value

    @staticmethod
    def _normalize_display_source_value(raw_value, attr_type: str):
        if attr_type not in {"enum", "user", "organization"} or not isinstance(raw_value, str):
            return raw_value

        stripped = raw_value.strip()
        if not stripped:
            return raw_value

        if stripped.startswith("[") and stripped.endswith("]"):
            try:
                parsed = ast.literal_eval(stripped)
            except (ValueError, SyntaxError):
                parsed = None
            if isinstance(parsed, (list, tuple, set)):
                if attr_type == "enum":
                    return normalize_enum_values(list(parsed))
                return list(parsed)

        return raw_value

    @staticmethod
    def _resolve_visible_columns(attr_map: dict[str, dict], show_fields: list[str] | None) -> list[str]:
        attr_ids = [attr_id for attr_id in attr_map.keys() if attr_id and not str(attr_id).endswith(DISPLAY_SUFFIX)]
        ordered: list[str] = []
        seen: set[str] = set()

        def append(key: str) -> None:
            if not key or key in seen:
                return
            seen.add(key)
            ordered.append(key)

        append("inst_name")
        if show_fields:
            for key in show_fields:
                if key == "inst_name" or key in attr_map:
                    append(key)
            return ordered

        for key in attr_ids:
            append(key)
        return ordered

    @staticmethod
    def _get_show_fields(model_id: str, user=None) -> list[str] | None:
        username = getattr(user, "username", None) if user is not None else None
        if not username:
            return None
        info = InstanceManage.get_info(model_id, username)
        if not isinstance(info, dict):
            return None
        fields = info.get("show_fields")
        if not isinstance(fields, list) or not fields:
            return None
        return [str(item) for item in fields if item]

    @staticmethod
    def _get_model_name(model_id: str) -> str:
        info = ModelManage.search_model_info(model_id) or {}
        if not isinstance(info, dict):
            return model_id
        return str(info.get("model_name") or model_id)

    @staticmethod
    def _build_model_attr_map(model_id: str) -> dict[str, dict]:
        attrs = ModelManage.search_model_attr_v2(model_id) or []
        return {str(attr.get("attr_id")): attr for attr in attrs if attr.get("attr_id") and not attr.get("is_display_field")}

    @classmethod
    def _get_cloud_region_name_map(cls) -> dict[str, str]:
        if cls._cloud_region_name_cache is not None:
            return cls._cloud_region_name_cache
        rows = NodeMgmt().cloud_region_list() or []
        result = {}
        for item in rows:
            if not isinstance(item, dict):
                continue
            region_id = item.get("id")
            region_name = item.get("name")
            if region_id in (None, "") or region_name in (None, ""):
                continue
            result[str(region_id)] = str(region_name)
        cls._cloud_region_name_cache = result
        return result

    @staticmethod
    def _get_display_value(instance: dict, field: str, attr_map: dict[str, dict]):
        if field == "cloud":
            cloud_map = ApplicationResourceOverviewService._get_cloud_region_name_map()
            raw_value = instance.get(field)
            if isinstance(raw_value, list):
                return ", ".join([cloud_map.get(str(item), str(item)) for item in raw_value if item not in (None, "")])
            if raw_value in (None, ""):
                return ""
            return cloud_map.get(str(raw_value), str(raw_value))

        attr = attr_map.get(field)
        if not attr:
            return ApplicationResourceOverviewService._serialize_instance_value(instance.get(field))

        attr_type = attr.get("attr_type")
        raw_value = ApplicationResourceOverviewService._normalize_display_source_value(instance.get(field), attr_type)
        if raw_value is not instance.get(field):
            instance[field] = raw_value
        display_field = f"{field}{DISPLAY_SUFFIX}"
        if attr_type in DISPLAY_FIELD_TYPES or is_file_attr_type(attr_type):
            if attr_type == "user" and not ApplicationResourceOverviewService._values_look_like_ids(raw_value):
                return ApplicationResourceOverviewService._serialize_instance_value(raw_value)
            DisplayFieldHandler.build_display_fields(
                attr.get("model_id") or instance.get("model_id") or "",
                instance,
                list(attr_map.values()),
            )
            display_value = instance.get(display_field)
            if display_value not in (None, ""):
                return ApplicationResourceOverviewService._serialize_instance_value(display_value)

        return ApplicationResourceOverviewService._serialize_instance_value(raw_value)

    @staticmethod
    def _values_look_like_ids(raw_value) -> bool:
        if raw_value in (None, ""):
            return False
        values = raw_value if isinstance(raw_value, (list, tuple, set)) else [raw_value]
        has_value = False
        for item in values:
            if item in (None, ""):
                continue
            has_value = True
            if isinstance(item, bool):
                return False
            if isinstance(item, int):
                continue
            if isinstance(item, str) and item.isdigit():
                continue
            return False
        return has_value

    @staticmethod
    def _build_instance_column_defs(columns: list[str], attr_map: dict[str, dict]) -> list[dict]:
        result = []
        for column in columns:
            attr = attr_map.get(column)
            title = (
                attr.get("attr_name")
                if attr and attr.get("attr_name")
                else ApplicationResourceOverviewService.INSTANCE_EXPORT_CORE_TITLES.get(column, column)
            )
            result.append({"key": column, "title": title})
        return result

    @staticmethod
    def build_topology_instance_groups(
        node_ids: list[str | int],
        permission_map: dict | None = None,
        user=None,
    ) -> dict:
        normalized_ids = []
        id_order = {}
        for raw_id in node_ids or []:
            try:
                inst_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if inst_id in id_order:
                continue
            id_order[inst_id] = len(normalized_ids)
            normalized_ids.append(inst_id)

        if not normalized_ids:
            return {"groups": [], "total": 0}

        instances = InstanceManage.query_entity_by_ids(normalized_ids) or []
        visible_instances = ApplicationResourceOverviewService._filter_visible_instances(
            instances,
            permission_map,
            user=user,
        )
        visible_instances.sort(key=lambda item: id_order.get(int(item.get("_id", 0)), 10**9))

        grouped = {}
        for instance in visible_instances:
            model_id = str(instance.get("model_id") or "unknown")
            group = grouped.setdefault(model_id, {"model_id": model_id, "columns": [], "items": [], "column_defs": []})
            group["items"].append(instance)

        groups = []
        for model_id, group in grouped.items():
            attr_map = ApplicationResourceOverviewService._build_model_attr_map(model_id)
            columns = ApplicationResourceOverviewService._resolve_visible_columns(
                attr_map,
                ApplicationResourceOverviewService._get_show_fields(model_id, user),
            )
            column_defs = ApplicationResourceOverviewService._build_instance_column_defs(columns, attr_map)
            identity = ApplicationResourceOverviewService.INSTANCE_IDENTITY_FIELDS
            groups.append(
                {
                    "model_id": model_id,
                    "model_name": ApplicationResourceOverviewService._get_model_name(model_id),
                    "columns": columns,
                    "column_defs": column_defs,
                    "count": len(group["items"]),
                    "items": [
                        {
                            **{key: ApplicationResourceOverviewService._get_display_value(instance, key, attr_map) for key in columns},
                            **{
                                key: ApplicationResourceOverviewService._serialize_instance_value(instance.get(key, ""))
                                for key in identity
                                if key not in columns
                            },
                        }
                        for instance in group["items"]
                    ],
                }
            )

        groups.sort(key=lambda item: item["model_id"])
        return {"groups": groups, "total": sum(item["count"] for item in groups)}

    @staticmethod
    def export_topology_instance_groups_excel(
        node_ids: list[str | int],
        permission_map: dict | None = None,
        user=None,
    ) -> bytes:
        payload = ApplicationResourceOverviewService.build_topology_instance_groups(
            node_ids=node_ids,
            permission_map=permission_map,
            user=user,
        )
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        groups = payload.get("groups") or []
        if not groups:
            sheet = workbook.create_sheet("instances")
            sheet.append(["message"])
            sheet.append(["no data"])
        else:
            used_names = set()
            for group in groups:
                raw_name = str(group.get("model_id") or "instances")
                safe_name = "".join("_" if ch in "[]:*?/\\\\" else ch for ch in raw_name)[:31] or "instances"
                candidate = safe_name
                suffix = 1
                while candidate in used_names:
                    suffix_text = f"_{suffix}"
                    candidate = f"{safe_name[:31 - len(suffix_text)]}{suffix_text}"
                    suffix += 1
                used_names.add(candidate)
                sheet = workbook.create_sheet(candidate)
                columns = group.get("columns") or []
                column_defs = group.get("column_defs") or []
                title_map = {item.get("key"): item.get("title") for item in column_defs}
                sheet.append([title_map.get(column, column) for column in columns])
                for item in group.get("items") or []:
                    sheet.append([item.get(column, "") for column in columns])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
