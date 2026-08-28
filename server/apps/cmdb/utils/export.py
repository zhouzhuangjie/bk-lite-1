from io import BytesIO
import json

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.cmdb.constants.constants import (
    ENUM,
    ORGANIZATION,
    USER,
    ASSOCIATION_TYPE,
    ATTR_TYPE_MAP,
)
from apps.cmdb.model_ops.extensions import is_file_attr_type
from apps.cmdb.services.model import ModelManage


def serialize_tag_values_for_export(values: list[str]) -> str:
    if not values:
        return ""
    return ",".join([str(value).strip() for value in values if str(value).strip()])


class Export:
    def __init__(self, attrs, model_id: str = "", association: list = None):
        self.attrs = attrs
        self.model_id = model_id
        self.association = association if association is not None else []
        self.association_type_map = {}
        self.model_name_map = {}
        self.model_asso_id_map = {}
        if self.association:
            self.association_type_map = {
                i["asst_id"]: i["asst_name"] for i in ASSOCIATION_TYPE
            }
            self.set_model_name_map()

    @staticmethod
    def _format_user_display_username(user_option: dict | None):
        """将用户 option(dict) 格式化为 display_name(username)，返回None表示无法格式化。"""
        if not user_option:
            return None

        username = user_option.get("username") or user_option.get("name")
        display_name = user_option.get("display_name")
        if display_name and username:
            return f"{display_name}({username})"
        return display_name or username

    def set_model_name_map(self):
        models = ModelManage.search_model()
        for model in models:
            self.model_name_map[model["model_id"]] = model["model_name"]

    def set_row_color(self, sheet, row_num, color):
        """行添加颜色"""
        for cell in sheet[row_num]:
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    def set_cell_color(self, sheet, row, col, color):
        """给指定单元格添加颜色"""
        cell = sheet.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def generate_header(self):
        """创建Excel文件, 设置属性与样式"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # 设置sheet名称为model_id
        sheet.title = self.model_id
        sheet.sheet_format.defaultColWidth = 20
        sheet.sheet_format.defaultRowHeight = 15
        attrs_name, attrs_type, attrs_id, index = (
            ["字段名(请勿编辑)"],
            ["字段类型(请勿编辑)"],
            ["字段标识(请勿编辑)"],
            0,
        )

        for attr_info in self.attrs:
            # 过滤掉 _display 冗余字段
            if attr_info.get("is_display_field"):
                continue
            # 附件/图片字段（企业版）不进入 Excel 导入导出
            if is_file_attr_type(attr_info.get("attr_type")):
                continue
            attr_name = (
                f"{attr_info['attr_name']}(必填)"
                if attr_info.get("is_required")
                else attr_info["attr_name"]
            )
            attrs_name.append(attr_name)
            attrs_id.append(attr_info["attr_id"])
            index += 1
            if attr_info["attr_type"] in {ENUM}:
                # 修复：Excel列索引需要+1，因为第一列是"字段名(请勿编辑)"
                col_index = index + 1
                sheet.add_data_validation(
                    self.set_enum_validation_by_sheet_data(
                        workbook, attr_info["attr_name"], attr_info["option"], col_index
                    )
                )
            attrs_type.append(ATTR_TYPE_MAP[attr_info["attr_type"]])

        for association in self.association:
            asst_id = association["asst_id"]
            dst_model_id = association["dst_model_id"]
            src_model_id = association["src_model_id"]
            model_asst_id = association["model_asst_id"]
            related_model_id = (
                src_model_id if self.model_id == dst_model_id else dst_model_id
            )
            _asst_model = self.model_name_map.get(related_model_id)
            if not _asst_model:
                continue
            asso_name = f"关联-{self.association_type_map[asst_id]}-{_asst_model}"
            attrs_name.append(asso_name)
            attrs_type.append("关联")
            attrs_id.append(model_asst_id)
            self.model_asso_id_map[model_asst_id] = {_asst_model: model_asst_id}

        sheet.append(attrs_name)
        sheet.append(attrs_type)
        sheet.append(attrs_id)
        self.set_row_color(sheet, 1, "92D050")
        self.set_row_color(sheet, 2, "C6EFCE")
        self.set_row_color(sheet, 3, "C6EFCE")
        # 给第一第二第三行的第一列添加橘黄色
        self.set_cell_color(sheet, 1, 1, "FFA500")  # A1
        self.set_cell_color(sheet, 2, 1, "FFA500")  # A2
        self.set_cell_color(sheet, 3, 1, "FFA500")  # A3

        return workbook

    def return_bytesio(self, workbook):
        """返回一个文件流"""
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return file_stream

    def set_enum_validation_by_sheet_data(self, workbook, filed_name, option, index):
        """设置枚举值, 通过sheet数据, 单选"""
        value_list = [i["name"] for i in option]

        # 将枚举数据放入sheet页
        filed_sheet = workbook.create_sheet(title=filed_name)
        for r, v in enumerate(value_list, start=1):
            filed_sheet.cell(row=r, column=1, value=v)

        # 创建 DataValidation 对象
        col = get_column_letter(index)
        last_row = len(filed_sheet["A"])
        dv = DataValidation(
            type="list", formula1=f"='{filed_sheet.title}'!$A$1:$A{last_row}"
        )
        dv.sqref = f"{col}4:{col}999"

        return dv

    def export_template(self):
        """导出模板"""
        workbook = self.generate_header()
        return self.return_bytesio(workbook)

    def export_inst_list(self, inst_list):
        """导出实例列表"""
        workbook = self.generate_header()
        # 找出枚举属性(过滤掉 _display 字段)
        enum_field_dict = {
            attr_info["attr_id"]: {i["id"]: i["name"] for i in attr_info["option"]}
            for attr_info in self.attrs
            if attr_info["attr_type"] in {ORGANIZATION, USER, ENUM}
            and not attr_info.get("is_display_field")
        }
        user_option_dict = {
            attr_info["attr_id"]: {i.get("id"): i for i in attr_info.get("option", [])}
            for attr_info in self.attrs
            if attr_info["attr_type"] == USER and not attr_info.get("is_display_field")
        }
        for inst_info in inst_list:
            sheet_data = [""]
            for attr in self.attrs:
                # 过滤掉 _display 冗余字段
                if attr.get("is_display_field"):
                    continue
                if attr["attr_type"] in {ORGANIZATION, USER}:
                    # attr_id_value = inst_info.get(attr["attr_id"], [])
                    # if not isinstance(attr_id_value, list):
                    #     attr_id_value = [attr_id_value]
                    # sheet_data.append(
                    #     str([enum_field_dict[attr["attr_id"]].get(i) for i in attr_id_value])
                    # )
                    attr_id_value = inst_info.get(attr["attr_id"], "")
                    # 主要维护人字段（operator）：支持多值，并格式化为 display_name(username)
                    if attr["attr_type"] == USER and attr.get("attr_id") == "operator":
                        if isinstance(attr_id_value, list):
                            formatted = []
                            for uid in attr_id_value:
                                text = self._format_user_display_username(
                                    user_option_dict.get(attr["attr_id"], {}).get(uid)
                                )
                                if text:
                                    formatted.append(text)
                                else:
                                    mapped = enum_field_dict.get(
                                        attr["attr_id"], {}
                                    ).get(uid)
                                    if mapped is not None:
                                        formatted.append(str(mapped))
                                    elif uid not in (None, ""):
                                        formatted.append(str(uid))
                            sheet_data.append(",".join(formatted))
                        else:
                            text = self._format_user_display_username(
                                user_option_dict.get(attr["attr_id"], {}).get(
                                    attr_id_value
                                )
                            )
                            if text:
                                sheet_data.append(text)
                            else:
                                mapped = enum_field_dict.get(attr["attr_id"], {}).get(
                                    attr_id_value
                                )
                                sheet_data.append(
                                    str(mapped) if mapped is not None else ""
                                )
                        continue

                    # 其他组织/用户字段保持原有导出格式
                    # TODO 目前只支持单选组织和用户，所以导出返回str即可 若支持单选则返回[]
                    if isinstance(attr_id_value, list):
                        if len(attr_id_value) > 0:
                            name = ",".join(
                                [
                                    str(enum_field_dict[attr["attr_id"]].get(i))
                                    for i in attr_id_value
                                ]
                            )
                            sheet_data.append(name)
                        else:
                            # 兼容空列表，避免 dict.get(list) 触发 TypeError 导致导出 500
                            sheet_data.append("")
                    else:
                        sheet_data.append(
                            str(enum_field_dict[attr["attr_id"]].get(attr_id_value))
                        )
                    continue

                if attr["attr_type"] == "tag":
                    tag_values = inst_info.get(attr["attr_id"], [])
                    if isinstance(tag_values, list):
                        sheet_data.append(serialize_tag_values_for_export(tag_values))
                    elif isinstance(tag_values, str):
                        sheet_data.append(tag_values)
                    else:
                        sheet_data.append("")
                    continue

                _value = inst_info.get(attr["attr_id"])
                if attr["attr_type"] == ENUM:
                    if isinstance(_value, list):
                        names = [
                            str(enum_field_dict[attr["attr_id"]].get(v, v))
                            for v in _value
                            if v is not None
                        ]
                        _value = ",".join(names)
                    else:
                        _value = enum_field_dict[attr["attr_id"]].get(_value)
                elif attr["attr_type"] == "table":
                    # table字段导出为单列JSON字符串
                    if _value:
                        if isinstance(_value, str):
                            # 已经是JSON字符串,直接使用
                            pass
                        else:
                            # 如果是列表/字典,序列化为JSON
                            _value = json.dumps(_value, ensure_ascii=False)
                    else:
                        _value = ""
                sheet_data.append(_value)
            # 查询当前实例的全部关联关系数据
            self.format_inst_asst_name(inst_info, sheet_data)
            workbook.active.append(sheet_data)
        return self.return_bytesio(workbook)

    def format_inst_asst_name(self, inst_info, sheet_data):
        from apps.cmdb.services.instance import InstanceManage

        inst_id = inst_info["_id"]

        # 只有在用户选择了关联关系时才查询，避免不必要的关联数据导出
        if not self.association:
            return

        # 获取所有关联关系数据
        asso_insts = InstanceManage.instance_association_instance_list(
            self.model_id,
            int(inst_id),
            business_only=True,
        )
        model_asst_name_map = {}
        for asso_inst in asso_insts:
            model_asst_id = asso_inst["model_asst_id"]
            model_asst_name_map[model_asst_id] = [
                inst["inst_name"] for inst in asso_inst["inst_list"]
            ]

        # 只处理用户选择的关联关系，而不是所有关联关系
        for association in self.association:
            model_asst_id = association["model_asst_id"]
            if model_asst_id in model_asst_name_map:
                data = ",".join(model_asst_name_map[model_asst_id])
            else:
                data = ""
            sheet_data.append(data)
